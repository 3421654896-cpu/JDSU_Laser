import sys
import serial
import math
import time
import yaml
import pyvisa
import openpyxl
import threading
import statistics

import pyqtgraph as pg
import numpy as np

from openpyxl import load_workbook
from PyQt5 import QtWidgets, QtCore, QtGui
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget,QLineEdit,QPushButton,
    QVBoxLayout, QLabel, QStackedWidget, QAction, QMessageBox,
    QFileDialog, QPlainTextEdit, QRadioButton, QButtonGroup
)
from PyQt5.QtCore import (
    QThread, pyqtSignal
)
from serial.tools import list_ports
from pathlib import Path
from datetime import datetime
from collections import deque
from queue import Queue
from queue import Empty

from scipy.signal import medfilt, butter, filtfilt
from pyvisa.constants import InterfaceType

#测试
# ====== 参数（按需改）======
ADDR = "GPIB0::7::INSTR"

TARGET_VENDOR = "YOKOGAWA"
TARGET_MODEL = "AQ6150"
TARGET_SN = "91P102177"

PRINT_EXCEL_EVERY = 1      # 每隔N行打印一次Excel数据（1=每行都打印）
PRINT_ACK = True           # 打印接收到的ACK

VISA_TIMEOUT_MS = 3000
VISA_RETRY = 10

FLUSH_EVERY_N = 1000

ACK_VALUE = 0x21
ACK_RESEND_SLEEP_S = 0.001  # 每次重发后短暂停一下，避免占满CPU

array_size = 4000

tx_size = 20
# ==========================

ser = serial.Serial(timeout=0.2)

ser_open = False
ser_cond = threading.Condition()

ap_open = False
ap_cond = threading.Condition()

switch_mode_enable = True
dac_type = 0

rx_queue = deque(maxlen=4000)

rx_buffer = bytearray()

frames_queue = deque(maxlen=200)

def get_desktop_path():
    return str(Path.home() / "Desktop")

def serial_write(info: bytes):
    ser.write(info)

def try_write(inst, cmd):
    try:
        inst.write(cmd)
        return True
    except Exception:
        return False

def parse_arr(reply: str):
    if reply is None:
        return []
    s = str(reply).strip().replace("\r", "").replace("\n", "")
    if not s:
        return []
    out = []
    for p in s.split(","):
        p = p.strip()
        if not p:
            continue
        try:
            out.append(float(p))
        except Exception:
            pass
    return out

def trunc6(x):
    """mW 建议保留 6 位小数更有意义"""
    if x is None or (isinstance(x, float) and math.isnan(x)):
        return "nan"
    try:
        return f"{float(x):.6f}"
    except Exception:
        return "nan"

def trunc4(x):
    """mW 建议保留 6 位小数更有意义"""
    if x is None or (isinstance(x, float) and math.isnan(x)):
        return "nan"
    try:
        return f"{float(x):.4f}"
    except Exception:
        return "nan"

def trunc3(x):
    if x is None or (isinstance(x, float) and math.isnan(x)):
        return "nan"
    try:
        return f"{math.trunc(float(x) * 1000) / 1000.0:.3f}"
    except Exception:
        return "nan"

def dbm_to_mw(dbm: float) -> float:
    # mW = 10^(dBm/10)
    return 10 ** (dbm / 10.0)

def get_top2_peaks_from_arrays(wav_list, pow_list_dbm):
    n = min(len(wav_list), len(pow_list_dbm))
    wav_list = wav_list[:n]
    pow_list_dbm = pow_list_dbm[:n]

    def is_small_int(x):
        return isinstance(x, (int, float)) and abs(x - round(x)) < 1e-9 and 0 <= x <= 200

    # 去掉头部计数 N
    if n >= 2 and is_small_int(wav_list[0]) and is_small_int(pow_list_dbm[0]):
        w1 = wav_list[1]
        if (500 <= w1 <= 20000) or (1e-9 <= w1 <= 1e-3):
            wav_list = wav_list[1:]
            pow_list_dbm = pow_list_dbm[1:]
            n -= 1

    # 单位判断：m->nm / nm直接用
    wav_nm = []
    for w in wav_list:
        w = float(w)
        if 1e-9 <= w <= 1e-3:
            wav_nm.append(w * 1e9)
        else:
            wav_nm.append(w)

    valid = [i for i in range(n) if 500 <= wav_nm[i] <= 20000]
    if not valid:
        return float("nan"), float("nan"), float("nan"), float("nan")

    valid.sort(key=lambda i: float(pow_list_dbm[i]), reverse=True)

    i1 = valid[0]
    wav1_nm = wav_nm[i1]
    pow1_mw = dbm_to_mw(float(pow_list_dbm[i1]))

    wav2_nm = pow2_mw = float("nan")
    if len(valid) >= 2:
        i2 = valid[1]
        wav2_nm = wav_nm[i2]
        pow2_mw = dbm_to_mw(float(pow_list_dbm[i2]))

    return wav1_nm, pow1_mw, wav2_nm, pow2_mw

def read_two_peaks_stable(inst):
    inst.write(":INIT")
    inst.write("*TRG")
    wav_reply = inst.query(":FETC:ARR:POW:WAV?")
    pow_reply = inst.query(":FETC:ARR:POW?")  # dBm
    wav_list = parse_arr(wav_reply)
    pow_list_dbm = parse_arr(pow_reply)
    return get_top2_peaks_from_arrays(wav_list, pow_list_dbm)

class MQComboBox(QtWidgets.QComboBox):
    def __init__(self):
        super().__init__()
        self.refresh_ports()

    def showPopup(self):
        self.refresh_ports()
        super().showPopup()

    def refresh_ports(self):
        current_text = self.currentText()

        self.blockSignals(True)
        self.clear()

        for p in list_ports.comports():
            text = f"{p.device} {p.description}"
            self.addItem(text, p.device)

        if current_text:
            index = self.findText(current_text)
            if index >= 0:
                self.setCurrentIndex(index)
        
        self.blockSignals(False)

class LogWidget(QPlainTextEdit):
    def __init__(self, max_lines=1000, readOnly=True):
        super().__init__()

        self.setReadOnly(readOnly)
        self.max_lines = max_lines

        self.setStyleSheet("""
        QPlainTextEdit{
            background-color: #1e1e1e;
            color: #dddddd;
            font-family: Consolas;
            font-size: 12px
        }
        """)

    def log(self, message, level="INFO"):
        time_str = datetime.now().strftime("%H:%M:%S")
        
        if level == "ERROR":
            text = f'<span style="color:#ff5555">[{time_str}] {message}</span>'
        elif level == "WARNING":
            text = f'<span style="color:#ffaa00">[{time_str}] {message}</span>'
        else:
            text = f'<span style="color:#dddddd">[{time_str}] {message}</span>'

        self.appendHtml(text)

        scrollbar = self.verticalScrollBar()
        scrollbar.setValue(scrollbar.maximum())

        if self.blockCount() > self.max_lines:
            cursor = self.textCursor()
            cursor.movePosition(cursor.Start)
            cursor.select(cursor.LineUnderCursor)
            cursor.removeSelectedText()
            cursor.deleteChar()

class peakWorker(QThread):
    temp_signal = pyqtSignal(float)

    def __init__(self):
        super().__init__()
        self.running = True
        self.perx_size = 20

    def run(self):
        pack_size = 4+2500*12+2+60*4+5+4 # 帧长度是变长的，但是接收的时候按定长为标准，接收不定长的帧
        def process_buffer(buf):
            while self.running:
                idx = buf.find(b'\xee\xee')
                if idx<0:
                    if len(buf) > 1:
                        del buf[:-1]
                        break
                elif idx>0:
                    # print(buf)
                    del buf[:idx]
                
                end_idx = buf.find(b'\xff\xef')
                if end_idx<0:
                    break

                frame = buf[idx:end_idx+2]
                # print(frame)

                if frame[-2:] == b'\xff\xef':
                    frames_queue.append(frame)
                    del buf[:end_idx+2]

        def process_temperature(buf):
            while self.running:
                idx = buf.find(b'\xff\xff')
                if idx<0:
                    break

                temp_frame = buf[idx:idx+self.perx_size]
                if len(temp_frame)<self.perx_size:
                    break
                del buf[idx:idx+self.perx_size]

                temp_int_high = temp_frame[4]
                temp_int_low = temp_frame[5]
                temp_dec_high = temp_frame[6]
                temp_dec_low = temp_frame[7]
                temperature = (temp_int_high<<8) + temp_int_low + \
                                ((temp_dec_high<<8)+temp_dec_low)*0.0001
                self.temp_signal.emit(temperature)

        while self.running:
            with ser_cond:
                while not ser_open:
                    ser_cond.wait()
                    rx_buffer.clear()
            if ser.is_open:
                try:
                    data = ser.read(64)
                except Exception as e:
                    print("Serial Error:",e)
                    continue
                rx_buffer.extend(data)
                # process_temperature(rx_buffer)
                process_buffer(rx_buffer)

class APWorker(QThread):
    log_signal = pyqtSignal(str,str)
    temp_signal = pyqtSignal(float)

    def __init__(self, file_path):
        super().__init__()
        self.file_path = file_path
        self.aprx_size = 20
        self.running = True
        self.temperature = 0
        self.flag_queue = Queue()

    def log(self, msg, level="info"):
        self.log_signal.emit(msg, level)

    def excel_operate(self, iter_excel):
        """
        返回：(cmd_bytes, info_str)
        """
        try:
            row = next(iter_excel)
            readGain = int(row[6])
            readSOA = int(row[7])
            readPhase = int(row[8])
            readwaveA = int(row[9])
            readwaveB = int(row[10])

            bWrite = [0] * tx_size
            bWrite[0] = 0xFF
            bWrite[1] = 0xFF
            bWrite[2] = 0x00
            bWrite[3] = dac_type
            bWrite[4] = (readGain >> 8) & 0xFF
            bWrite[5] = (readGain >> 0) & 0xFF
            bWrite[6] = (readSOA >> 8) & 0xFF
            bWrite[7] = (readSOA >> 0) & 0xFF
            bWrite[8] = (readPhase >> 8) & 0xFF
            bWrite[9] = (readPhase >> 0) & 0xFF
            bWrite[10] = (readwaveA >> 8) & 0xFF
            bWrite[11] = (readwaveA >> 0) & 0xFF
            bWrite[12] = (readwaveB >> 8) & 0xFF
            bWrite[13] = (readwaveB >> 0) & 0xFF

            cmd = bytes(bWrite)
            info = f"Gain={readGain}, SOA={readSOA}, phase={readPhase}, waveA={readwaveA}, waveB={readwaveB}"
            return cmd, info
        except Exception:
            return None, None

    def serial_recv_loop(self):
        while self.running:
            b = list(ser.read(self.aprx_size))
            if not b:
                continue

            if b[0] == 0xFF and b[1] == 0xFF:
                if b[3] == 0x00 and b[4] == 0x21:
                    self.flag_queue.put(b)

                elif b[3] == 0x01:
                    temp_int_high = b[4]
                    temp_int_low = b[5]
                    temp_dec_high = b[6]
                    temp_dec_low = b[7]
                    self.temperature = (temp_int_high<<8) + temp_int_low + \
                                    ((temp_dec_high<<8)+temp_dec_low)*0.0001
                    self.temp_signal.emit(self.temperature)

    def connect_aq6150(self, rm):
        # 只看资源信息，不真正打开设备
        infos = rm.list_resources_info()

        # 只保留 GPIB INSTR，完全跳过 ASRL/USB/TCPIP
        gpib_resources = [
            name for name, info in infos.items()
            if info.interface_type == InterfaceType.gpib
            and info.resource_class == "INSTR"
        ]

        for res in gpib_resources:
            inst = None
            try:
                inst = rm.open_resource(res)
                inst.timeout = VISA_TIMEOUT_MS
                inst.read_termination = "\n"
                inst.write_termination = "\n"

                idn = inst.query("*IDN?").strip()
                parts = [x.strip() for x in idn.split(",")]

                if len(parts) >= 3:
                    vendor, model, serial = parts[:3]
                    if vendor == TARGET_VENDOR and model == TARGET_MODEL and serial == TARGET_SN:
                        self.log(f"找到 AQ6150: {res} -> {idn}")
                        return inst

                inst.close()

            except Exception:
                if inst is not None:
                    try:
                        inst.close()
                    except:
                        pass

        self.log("未找到目标 AQ6150", "ERROR")

    def run(self):
        try:
            recv_thread = threading.Thread(target=self.serial_recv_loop, daemon=True)
            recv_thread.start()

            XLSX_BASENAME = f"AQ6150B_log_{time.strftime('%H%M%S')}.xlsx"
            excel_path = self.file_path
            if not excel_path:
                self.log("未选择文件", "WARNING")
                return
            self.log(f"执行文件: {excel_path}")

            wb_in = load_workbook(excel_path, read_only=True, data_only=True)
            ws_in = wb_in.active
            iter_excel = ws_in.iter_rows(min_row=2, values_only=True)

            rm = pyvisa.ResourceManager()
            inst = self.connect_aq6150(rm)

            try:
                # 设置单位仍然用 dBm（我们在软件里转 mW）
                try_write(inst, ":INIT:CONT OFF")
                try_write(inst, ":TRIG:SOUR BUS")
                try_write(inst, ":FORM:NDAT 0NM")
                try_write(inst, ":UNIT:POW DBM")

                desktop = get_desktop_path()
                xlsx_path = Path(desktop) / XLSX_BASENAME
                # file_exists = os.path.isfile(xlsx_path)

                # with open(xlsx_path, "a", newline="", encoding="utf-8") as f:
                #     writer = csv.writer(f)
                wb_init = openpyxl.Workbook()
                wb_init.save(xlsx_path)
                wb_out = load_workbook(xlsx_path)
                ws_out = wb_out.active

                head_data = ["timestamp_iso",
                        "peak1_wavelength_nm", "peak1_power_mW",
                        "peak2_wavelength_nm", "peak2_power_mW", "temperature",
                        "pdr", "pdt", "ratio_rt", "r_sat_sa"]
                gap_len = len(head_data)+1

                self.log("开始记录(ACK 不成功则一直重发同一行，不推进 Excel)。Ctrl+C 结束。")

                count = 0
                since_flush = 0
                loop_time = -1

                # 缓存当前行（ACK 成功后才读取下一行）
                current_cmd = None
                current_info = None
                last_current = True

                while self.running:
                    # 循环写入表头
                    if last_current:
                        count=0
                        loop_time += 1
                        iter_excel = ws_in.iter_rows(min_row=2, values_only=True)
                        for i, value in enumerate(head_data):
                            ws_out.cell(row=1, column=i+1+loop_time*gap_len, value=value)
                        last_current = False
                    
                    # 只有当当前没有待发送行时，才取 Excel 下一行
                    if current_cmd is None:
                        cmd, info = self.excel_operate(iter_excel)
                        if cmd is None:
                            last_current = True
                            continue
                        current_cmd, current_info = cmd, info

                    # 打印 Excel 读取内容（可控频率）
                    if PRINT_EXCEL_EVERY > 0 and ((count + 1) % PRINT_EXCEL_EVERY == 0):
                        self.log("----excel----")
                        self.log(current_info)
                        self.log("----excel----")

                    while True:
                        try:
                            self.flag_queue.get_nowait()
                        except Empty:
                            break

                    # 串口发送
                    serial_write(current_cmd)

                    # 等 ACK（不成功就一直重发，不推进）
                    while self.running:
                        try:
                            flag = self.flag_queue.get(timeout=3.0)
                            break
                        except Empty:
                            serial_write(current_cmd)

                    # ACK 成功：计数并“推进到下一行”
                    count += 1
                    self.log(f"循环:{loop_time+1}  计数:{count}")
                    current_cmd = None
                    current_info = None

                    # 读 PDT和PDR
                    pdt = (flag[5]<<8)+flag[6]
                    pdr = (flag[8]<<8)+flag[9]

                    pdt_sa,pdr_sa = flag[7],flag[10]

                    pdt = pdt*2.5/4096
                    pdr = pdr*2.5/4096

                    pdt = (1.25-pdt)/2
                    pdr = (1.25-pdr)/2

                    # 读 AQ（失败短重试）
                    ts = datetime.now().isoformat(timespec="milliseconds")
                    wav1_nm = pow1_mw = wav2_nm = pow2_mw = float("nan")
                    for _ in range(VISA_RETRY + 1):
                        try:
                            wav1_nm, pow1_mw, wav2_nm, pow2_mw = read_two_peaks_stable(inst)
                            break
                        except Exception:
                            inst = self.connect_aq6150(rm)
                            time.sleep(1)

                    # 写（波长保留3位，功率mW保留6位）
                    cur_row = [ts, trunc3(wav1_nm), trunc6(pow1_mw), trunc3(wav2_nm), trunc6(pow2_mw), trunc4(self.temperature), trunc4(pdr), trunc4(pdt), trunc3(pdr/pdt), f"{pdr_sa}{pdt_sa}"]
                    for i, value in enumerate(cur_row):
                        ws_out.cell(row=count+1, column=i+1+loop_time*gap_len, value=value)

                    # 清缓冲到硬盘
                    since_flush += 1
                    if since_flush >= FLUSH_EVERY_N:
                        wb_out.save(xlsx_path)
                        since_flush = 0

            except KeyboardInterrupt:
                self.log("\n用户中断, 已停止.")
                wb_out.save(xlsx_path)
                wb_out.close()

            finally:
                try:
                    inst.close()
                finally:
                    rm.close()
                wb_out.save(xlsx_path)
                wb_out.close()
                self.log("结束.")

        except Exception as e:
            self.log(str(e), "ERROR")


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()

        self.setWindowTitle("可调谐激光器")
        self.resize(1400, 800)

        menubar = self.menuBar()

        self.menu_port = menubar.addMenu("端口")
        self.menu_baud = menubar.addMenu("波特率")
        self.menu_page = menubar.addMenu("工作模式")
        self.menu_dac = menubar.addMenu("DAC选择")

        self.baud_rates = [9600, 115200, 2000000, 3000000, 4000000, 6000000]
        self.baud = 2000000
        self.baud_act = None

        self.ports = []
        listp = list_ports.comports()
        for p in listp:
            self.ports.append([p.device, p.description])
        target_cp = "CH343"
        self.port = next((p.device for p in listp if target_cp in p.description), 
                        listp[0].device if listp else None)
        self.port_act = None

        self.MCU_mode = 0
        self.mode_act = None

        self.dac_act = None

        global ser
        ser.port = self.port
        ser.baudrate = self.baud
        
        self.stack = QStackedWidget()
        # self.stack.currentChanged.connect(self.on_switch_mode)
        
        self.page_peak = GraphWindow()
        self.page_ap6150 = ap6150bWindow()
        self.page_extra = extraWindow()

        self.stack.addWidget(self.page_peak)
        self.stack.addWidget(self.page_ap6150)
        self.stack.addWidget(self.page_extra)

        self.setCentralWidget(self.stack)

        self.init_menu()

    def init_menu(self):
        # 端口
        self.menu_port.aboutToShow.connect(self.update_ports_menu)

        # 波特率
        for baud in self.baud_rates:
            action = QAction(str(baud), self)
            action.setCheckable(True)
            action.setData(baud)
            action.triggered.connect(self.set_baudrate)
            self.menu_baud.addAction(action)
            if baud == self.baud:
                action.setChecked(True)
                self.baud_act = action

        # 工作模式
        work_mode = ["寻峰模式", "扫波长模式", "单值模式"]
        for i, mode_str in enumerate(work_mode):
            action = QAction(mode_str,self)
            action.setCheckable(True)
            action.setData(i)
            action.triggered.connect(self.set_page)
            self.menu_page.addAction(action)
            if i == self.MCU_mode:
                self.mode_act = action
                action.trigger()

        # DAC选择
        global dac_type
        dac_targets = ["U_DAC", "I_DAC"]
        for i, dac_target in enumerate(dac_targets):
            action = QAction(dac_target,self)
            action.setCheckable(True)
            action.setData(i)
            action.triggered.connect(self.set_dac)
            self.menu_dac.addAction(action)
            if i == dac_type:
                self.dac_act = action
                action.trigger()
                
    def update_ports_menu(self):
        menu = self.sender()

        menu.clear()
        self.ports.clear()

        for p in list_ports.comports():
            self.ports.append([p.device, p.description])
            action = QAction(f"{p.device} {p.description}", self)
            action.setCheckable(True)
            action.setData(p.device)
            action.triggered.connect(self.set_com_port)
            menu.addAction(action)
            if p.device == self.port:
                action.setChecked(True)
                self.port_act = action

    def set_dac(self):
        action = self.sender()
        if not switch_mode_enable:
            self.dac_act.setChecked(True)
            action.setChecked(False)
            QMessageBox.warning(self, "警告", "先停止当前页面工作流再切换DAC")
            return

        command_frame = bytearray(tx_size)
        command_frame[0] = 0xFF
        command_frame[1] = 0xFF
        command_frame[2] = 0x01
        command_frame[3] = 0x03
        command_frame[8] = (action.data()) & 0xFF

        try:
            if not ser.is_open:
                ser.open()
            ser.write(bytes(command_frame))
            ser.close()
        except Exception as e:
            self.dac_act.setChecked(True)
            action.setChecked(False)
            QMessageBox.warning(self, "警告", "改DAC类型前确保串口端口和波特率选对")
            QMessageBox.critical(self, "crash", f"发生异常:\n{str(e)}")
            return

        global dac_type
        dac_type = action.data()
        self.dac_act.setChecked(False)
        action.setChecked(True)
        self.dac_act = action
        
    def set_page(self):
        action = self.sender()
        if not switch_mode_enable:
            self.mode_act.setChecked(True)
            action.setChecked(False)
            QMessageBox.warning(self, "警告", "先停止当前页面工作流再切换模式")
            return

        command_frame = bytearray(tx_size)
        command_frame[0] = 0xFF
        command_frame[1] = 0xFF
        command_frame[2] = 0x01
        command_frame[3] = 0x02
        command_frame[8] = (action.data()) & 0xFF

        try:
            if not ser.is_open:
                ser.open()
            ser.write(bytes(command_frame))
            ser.close()
        except Exception as e:
            self.mode_act.setChecked(True)
            action.setChecked(False)
            QMessageBox.warning(self, "警告", "改工作模式前确保串口端口和波特率选对")
            QMessageBox.critical(self, "crash", f"发生异常:\n{str(e)}")
            return

        self.MCU_mode = action.data()
        self.stack.setCurrentIndex(self.MCU_mode)
        self.mode_act.setChecked(False)
        action.setChecked(True)
        self.mode_act = action
        if self.MCU_mode == 2:
            self.page_extra.monitor_btn.click()

    def set_com_port(self):
        global ser
        action = self.sender()
        if switch_mode_enable:
            self.port = action.data()
            ser.port = self.port
            self.port_act.setChecked(False)
            action.setChecked(True)
            self.port_act = action
        else:
            action.setChecked(False)
            QMessageBox.warning(self, "警告", "先停止当前页面工作流停止再修改端口")

    def set_baudrate(self):
        global ser
        action = self.sender()
        if switch_mode_enable:
            self.baud = action.data()
            ser.baudrate = self.baud
            self.baud_act.setChecked(False)
            action.setChecked(True)
            self.baud_act = action
        else:
            action.setChecked(False)
            QMessageBox.warning(self, "警告", "先停止当前页面工作流停止再修改波特率")

    def closeEvent(self, a0):
        self.page_extra.stop_recv_thread()
        return super().closeEvent(a0)

class GraphWindow(QtWidgets.QWidget):
    def __init__(self):
        super().__init__()
        self.paused = False
        self.process_down = True
        self.peaks_lines = [[] for _ in range(4)]
        self.visible_lines = [False for _ in range(4)]
        self.us_points = []
        self.impress_points = []

        layout = QtWidgets.QGridLayout()
        self.setLayout(layout)

        self.ctrl_panel = QtWidgets.QWidget()
        self.ctrl_layout = QtWidgets.QHBoxLayout()
        self.ctrl_layout.setContentsMargins(5,5,5,5)
        self.ctrl_layout.setSpacing(10)
        self.ctrl_panel.setLayout(self.ctrl_layout)
        self.ctrl_panel.setMaximumHeight(80)

        lab_interval = QtWidgets.QLabel("单数据间隔(us):")
        self.interval_text = QtWidgets.QLineEdit("1")
        self.interval_text.setMinimumWidth(100)
        self.interval_text.setMaximumHeight(25)

        lab_temperature = QtWidgets.QLabel("温度(℃):")
        self.temperature_text = QtWidgets.QLineEdit("0")
        self.temperature_text.setReadOnly(True)
        self.temperature_text.setMinimumWidth(100)
        self.temperature_text.setMaximumHeight(25)
        self.temperature = 0

        self.com_btn = QtWidgets.QPushButton("打开")
        self.com_btn.setMinimumWidth(100)

        self.clear_btn = QtWidgets.QPushButton("清空数据")
        self.clear_btn.setMinimumWidth(100)

        self.com_btn.clicked.connect(self.on_open_changed)
        self.interval_text.textChanged.connect(self.on_interval_changed)
        self.clear_btn.clicked.connect(self.on_clear_chart)

        self.ctrl_layout.addWidget(lab_interval)
        self.ctrl_layout.addWidget(self.interval_text)
        self.ctrl_layout.addSpacing(20)
        self.ctrl_layout.addStretch()

        self.ctrl_layout.addWidget(lab_temperature)
        self.ctrl_layout.addWidget(self.temperature_text)
        self.ctrl_layout.addSpacing(20)
        self.ctrl_layout.addStretch()

        self.ctrl_layout.addWidget(self.com_btn)
        self.ctrl_layout.addSpacing(20)
        self.ctrl_layout.addStretch()

        self.ctrl_layout.addWidget(self.clear_btn)
        
        layout.addWidget(self.ctrl_panel, 0, 0)

        self.plot1 = pg.PlotWidget()

        layout.addWidget(self.plot1, 1, 0)

        self.adc = [deque(maxlen=array_size) for _ in range(4)]
        self.data = [deque(maxlen=array_size) for _ in range(4)]
        self.filts = [np.array(list()) for _ in range(4)]
        self.usdata = [list() for _ in range(4)]

        self.waves = [[0 for _ in range(15)] for _ in range(4)]

        self.plot1.addLegend()
        self.color_list = ['yellow','green','blue','purple']
        self.curve1 = self.plot1.plot(pen=self.color_list[0], name='CH0')
        self.curve2 = self.plot1.plot(pen=self.color_list[1], name='CH1')
        self.curve3 = self.plot1.plot(pen=self.color_list[2], name='CH2')
        self.curve4 = self.plot1.plot(pen=self.color_list[3], name='CH3')

        self.num_panel = QtWidgets.QWidget()
        self.num_layout = QtWidgets.QGridLayout()
        self.num_panel.setLayout(self.num_layout)

        header_style = "font-weight:bold; font-size:14pt; padding:3pt;"

        self.check_boxs = []

        for r in range(4):
            cb = QtWidgets.QCheckBox()
            cb.setChecked(False)
            cb.stateChanged.connect(lambda state, i=r: self.toggle_line(i, state))
            self.check_boxs.append(cb)
            self.num_layout.addWidget(cb, r+1, 0)

        # 左侧表头（行：1~4）
        for r in range(4):
            lab = QtWidgets.QLabel(str(r+1))
            lab.setAlignment(QtCore.Qt.AlignCenter)
            lab.setStyleSheet(header_style)
            self.num_layout.addWidget(lab, r+1, 1)

        # 顶部表头（列：1~15）
        for c in range(15):
            lab = QtWidgets.QLabel(str(c+1))
            lab.setAlignment(QtCore.Qt.AlignCenter)
            lab.setStyleSheet(header_style)
            self.num_layout.addWidget(lab, 0, c+2)

        self.num_labels = [[None for _ in range(15)] for _ in range(4)]

        for r in range(4):
            for c in range(15):
                lab = QtWidgets.QLabel("0")
                lab.setAlignment(QtCore.Qt.AlignCenter)
                lab.setStyleSheet("font-size:14pt; padding:2pt;")
                self.num_labels[r][c] = lab
                self.num_layout.addWidget(lab, r+1, c+2)

        layout.addWidget(self.num_panel, 2, 0, 1, 2)

        self.interval = int(self.interval_text.text())

        self.worker = peakWorker()
        self.worker.temp_signal.connect(self.update_temp)
        # self.worker.start()
        
        self.frame_timer = QtCore.QTimer()    
        self.frame_timer.timeout.connect(self.process_frame)
        
        self.update_timer = QtCore.QTimer()    
        self.update_timer.timeout.connect(self.update_plot)

        with open("./wave_const.yaml", 'r', encoding="utf-8") as file:
            yaml_data = yaml.safe_load(file)
            # print((yaml_data['Wave_DATA']))
            self.yaml = yaml_data['Wave_DATA']
            self.wave_const = [num[0]+num[1]*0.001 for num in self.yaml]
            self.plot1.setXRange(int(min(self.wave_const)), math.ceil(max(self.wave_const)))

        x_extend = 5
        self.voltage_range = 5
        self.plot1.setYRange(-0.5,2.5)
        self.plot1.getViewBox().setLimits(xMin=self.wave_const[0]-x_extend,xMax=self.wave_const[-1]+x_extend,
                                          yMin=-self.voltage_range/self.voltage_range,yMax=self.voltage_range)
        self.plot1.showGrid(x=True, y=True)

        self.visual_index = 0
        self.visual_y = 0
        self.vLine = pg.InfiniteLine(angle=90, movable=False, pen=pg.mkPen('r'))
        self.hLine = pg.InfiniteLine(angle=0, movable=False, pen=pg.mkPen('r'))
        self.plot1.addItem(self.vLine, ignoreBounds=True)
        self.plot1.addItem(self.hLine, ignoreBounds=True)

        self.label = pg.TextItem(color='w')
        self.plot1.addItem(self.label)

        self.plot1.proxy = pg.SignalProxy(
            self.plot1.scene().sigMouseMoved,
            rateLimit=60,
            slot=self.mouseMoved
        )

        self.initials_length = 15
        self.peak_interval = 30
        self.peak_threshold = 164

    def mouseMoved(self, evt):
        pos = evt[0]
        if self.plot1.sceneBoundingRect().contains(pos):

            mousePoint = self.plot1.plotItem.vb.mapSceneToView(pos)
            x = mousePoint.x()
            y = mousePoint.y()

            # 找最近的x索引
            index = np.abs(np.array(self.wave_const)-x).argmin()
            if index>=array_size or index<0 or len(self.adc[0])==0:
                return
            
            self.visual_index = index
            self.visual_y = y
            self.update_crosshair(index, y)

    def update_crosshair(self, index, y):
        if index >= array_size or index < 0 \
        or y > self.voltage_range or y < -self.voltage_range \
        or len(self.adc[0]) == 0:
            return
    
        adc = np.array([
            self.filts[0][index],
            self.filts[1][index],
            self.filts[2][index],
            self.filts[3][index]
        ])

        indey = np.abs(adc-y).argmin()

        x_snap = self.wave_const[index]
        y_snap = adc[indey]

        self.vLine.setPos(x_snap)
        self.hLine.setPos(y_snap)

        self.label.setText(f"x={x_snap:.3f}\ny={y_snap:.3f}")
        self.label.setPos(x_snap, y_snap)

    def on_open_changed(self):
        global ser_open
        global ser_cond
        global ser
        global switch_mode_enable

        do_open = False
        do_close = False

        with ser_cond: 
            if ser_open == False:
                do_open = True
                ser_open = True 
                ser_cond.notify_all() 
                
            else: 
                do_close = True
                ser_open = False

        if do_open:
            try:
                ser.open()
            except Exception as e:
                print("Serial Error:",e)
                with ser_cond:
                    ser_open = False
                return
            
            switch_mode_enable = False
            self.clear_btn.setEnabled(False)
            self.on_interval_changed()

            self.worker = peakWorker()
            self.worker.temp_signal.connect(self.update_temp)
            self.worker.start()

            self.frame_timer.start(1)
            self.update_timer.start(2)

            self.com_btn.setText("关闭")
        elif do_close:
            self.worker.running = False

            try:
                ser.close()
            except Exception as e:
                print("Serial Error:",e)
                with ser_cond:
                    ser_open = True
                return
            
            switch_mode_enable = True
            self.clear_btn.setEnabled(True)

            self.frame_timer.stop()
            self.update_timer.stop()

            # self.worker.quit()
            self.worker.wait()

            self.com_btn.setText("打开")

    def on_clear_chart(self):
        curves = [self.curve1, self.curve2, self.curve3, self.curve4]
        datas = self.data
        adcs = self.data
        for idx in range(4):
            curves[idx].setData([])
            datas[idx].clear()
            adcs[idx].clear()
            self.temperature = 0

            for j in range(len(self.peaks_lines[idx])):
                self.peaks_lines[idx][j].setVisible(False)
                del self.peaks_lines[idx][j]
            for label in self.num_labels[idx]:  
                label.setText("0")
            self.temperature_text.setText("0")

    def on_interval_changed(self):
        self.interval = int(self.interval_text.text())
        command_frame = [0]*tx_size
        command_frame[0] = 0xFF
        command_frame[1] = 0xFF
        command_frame[2] = 0x01
        command_frame[3] = 0x01
        command_frame[7] = (self.interval>>24) & 0xFF
        command_frame[8] = (self.interval>>16) & 0xFF
        command_frame[9] = (self.interval>>8) & 0xFF
        command_frame[10] = self.interval & 0xFF

        if not ser.is_open:
            try:
                self.com_btn.setEnabled(False)
                ser.open()
                ser.write(bytes(command_frame))
                ser.close()
                self.com_btn.setEnabled(True)
                QMessageBox.information(self, "提示", "间隔已生效")
            except Exception as e:
                QMessageBox.information(self, "提示", "改间隔前确保串口端口和波特率选对")
                QMessageBox.information(self, "警告", f"发生异常:\n{str(e)}")
                return
        else:
            ser.write(bytes(command_frame))

    def process_frame(self):
        try:
            if len(frames_queue)==0 or self.process_down==False:
                # print("no valid frame")
                return
            raw = frames_queue.popleft()
            self.process_down = False
            single_size = 12

            # print(len(raw))
            if raw[0]!=0xEE and raw[1]!=0xEE:
                print("pack head error")
                self.process_down = True
                # print(raw)
                return 
            
            global array_size
            _array_size = (raw[2]<<8)+raw[3]
            if not _array_size == array_size:
                array_size = _array_size
                # print(array_size)

                self.adc = [deque(maxlen=array_size) for _ in range(4)]
                self.data = [deque(maxlen=array_size) for _ in range(4)]

            self.usdata = [list() for _ in range(4)]

            com_index = 0
            for i in range(4, array_size*12+4, single_size):
                com_input = raw[i:i+single_size]

                ch1 = (com_input[0]<<8)+com_input[1]
                st1 = com_input[2]
                ch2 = (com_input[3]<<8)+com_input[4]
                st2 = com_input[5]
                ch3 = (com_input[6]<<8)+com_input[7]
                st3 = com_input[8]
                ch4 = (com_input[9]<<8)+com_input[10]
                st4 = com_input[11]

                v1 = ch1*2.5/4095
                v2 = ch2*2.5/4095
                v3 = ch3*2.5/4095
                v4 = ch4*2.5/4095

                self.adc[0].append(v1)
                self.adc[1].append(v2)
                self.adc[2].append(v3)
                self.adc[3].append(v4)

                self.data[0].append(ch1)
                self.data[1].append(ch2)
                self.data[2].append(ch3)
                self.data[3].append(ch4)

                if not st1 ==1:
                    self.usdata[0].append(com_index)
                if not st2 ==1:
                    self.usdata[1].append(com_index)
                if not st3 ==1:
                    self.usdata[2].append(com_index)
                if not st4 ==1:
                    self.usdata[3].append(com_index)

                com_index+=1

            if raw[array_size*12+4]!=0xAB:
                print("wave head error")
                print(hex(raw[array_size*12+4]))
                self.process_down = True
                return
            self.waves = [[0 for _ in range(15)] for _ in range(4)]
            frame_count = array_size*12+4
            data_len = 0
            for i in range(0, 4):
                frame_count += 1

                data_len = raw[frame_count]

                for j in range(data_len):
                    frame_count+=1
                    int_high = raw[frame_count]
                    frame_count+=1
                    int_low = raw[frame_count]
                    frame_count+=1
                    dec_high = raw[frame_count]
                    frame_count+=1
                    dec_low = raw[frame_count]
                    get_wave = (int_high<<8)+int_low+((dec_high<<8)+dec_low)*0.001

                    self.waves[i][j] = get_wave
            
            frame_count+=1
            temp_int_high = raw[frame_count]
            frame_count+=1
            temp_int_low = raw[frame_count]
            frame_count+=1
            temp_dec_high = raw[frame_count]
            frame_count+=1
            temp_dec_low = raw[frame_count]
            temperature = (temp_int_high<<8)+temp_int_low+((temp_dec_high<<8)+temp_dec_low)*0.0001
            self.update_temp(temperature)

            self.process_down = True
        except Exception as e:
            print(e)
            print(len(raw))
            self.process_down = True
            return

    def update_temp(self, _temperature):
        self.temperature = _temperature
        self.temperature_text.setText(f"{_temperature}")

    def update_us_point(self):
        us_list = self.usdata
        filt_list = self.filts
        x_data = []
        y_data = []
        for usp in self.us_points:
            self.plot1.removeItem(usp)
        self.us_points.clear()
        for usl,adcl in zip(us_list,filt_list):
            for pt in usl:
                x_data.append(self.wave_const[pt])
                y_data.append(adcl[pt])
                
            scatter = pg.ScatterPlotItem(x_data, y_data, pen=pg.mkPen(255,0,0), symbol='o', symbolBrush=pg.mkBrush(255,0,0), symbolSize=10)
            
            self.us_points.append(scatter)
            self.plot1.addItem(scatter)

    def update_plot(self):
        if self.process_down == False:
            return
        if not len(self.wave_const) == len(self.adc[0]):
            # print("size error")
            return
        self.process_down = False

        self.plot1.getViewBox().setLimits(xMin=self.wave_const[0]-5,xMax=self.wave_const[-1]+5,
                                          yMin=-self.voltage_range/self.voltage_range,yMax=self.voltage_range)
        x = self.wave_const

        for i in range(4):
            self.calculate_peaks(i)

        # 更新被抑制的点
        for imp in self.impress_points:
            self.plot1.removeItem(imp)

        # 更新曲线
        self.filts = [self.adc_filter(_adcs) for _adcs in self.adc]
        self.curve1.setData(np.array(x),np.array(self.filts[0]))
        self.curve2.setData(np.array(x),np.array(self.filts[1]))
        self.curve3.setData(np.array(x),np.array(self.filts[2]))
        self.curve4.setData(np.array(x),np.array(self.filts[3]))

        # 更新不稳定点
        self.update_us_point()

        # 更新光标
        self.update_crosshair(self.visual_index, self.visual_y)

        # 更新峰值可视化线
        for i in range(4):
            for j in range(15):
                self.num_labels[i][j].setText("{:.3f}".format(self.waves[i][j]) if not self.waves[i][j]==0 else "0")
            if self.visible_lines[i]:
                self.cal_peaks_line(i)
            else:
                for l in self.peaks_lines[i]:
                    l.setVisible(False)

        self.process_down = True

    def calculate_peaks(self, i):
        _data = self.data[i]
        if len(_data) == 0:
            return
        peaks = self.find_peaks(_data, i)
        # print(peaks)

    def cal_peaks_line(self, i):
        if len(self.data[i]) == 0:
            return
        for l in self.peaks_lines[i]:
            l.setVisible(False)

        for idx,p in enumerate(self.waves[i]):
            if p==0:
                if idx < len(self.peaks_lines[i]):
                    self.plot1.removeItem(self.peaks_lines[i][idx])
                    del self.peaks_lines[i][idx]
                continue
            if idx>=len(self.peaks_lines[i]):
                vline = pg.InfiniteLine(
                    pos=p,
                    angle=90,
                    pen=pg.mkPen('r',width=1,style=QtCore.Qt.DashLine)
                )
                self.plot1.addItem(vline)
                self.peaks_lines[i].append(vline)
            else:
                self.peaks_lines[i][idx].setPos(p)
                self.peaks_lines[i][idx].setVisible(True)

    def toggle_line(self, i, state):
        visible = (state == QtCore.Qt.Checked)
        self.visible_lines[i] = visible

    def find_initial(self, data_vec, initials, adc_length, adc_index):
        ma = max(data_vec)
        mi = min(data_vec)
        gap = ma-mi
        data_norvec = list(np.array(data_vec)-mi)
        mean = sum(data_norvec)
        # print(f"{adc_index}:", mean)
        
        if 100*mean>10*adc_length*gap:
            return initials
        it = 0
        for i in range(1,adc_length-2):
            if it>=self.initials_length: break
            start = i-self.peak_interval if i-self.peak_interval>=0 else 0
            end = i+self.peak_interval if i+self.peak_interval<adc_length else adc_length

            # print(data_norvec)
            front = list(data_norvec)[start:i]
            back = list(data_norvec)[i+1:end]
            # print(front)
            # print(back)
            if data_norvec[i]>=max(front) and data_norvec[i]>=max(back) and data_norvec[i]>self.peak_threshold:
                if 10*(data_norvec[i]-data_norvec[i-1])>gap*5 or 10*(data_norvec[i]-data_norvec[i-1])>gap*5:
                   self.adc[adc_index][i] = 0
                   continue
                # print(data_norvec[i])
                initials[it] = i
                i+=self.peak_interval
                it+=1
        # print(initials)
        return initials

    def find_peaks(self, data_vec, adc_index):
        initials = [0]*self.initials_length
        peaks_vec = [0]*self.initials_length
        adc_length = array_size

        initials = self.find_initial(data_vec, initials, adc_length, adc_index)
        
        for i in range(self.initials_length):
            if initials[i]==0:
                break
            start = initials[i]-self.peak_interval if initials[i]-self.peak_interval>=0 else 0
            end = initials[i]+self.peak_interval if initials[i]+self.peak_interval<adc_length else adc_length
            sumy=0
            sumxy=0
            for j in range(start, end):
                sumxy+=data_vec[j]*self.wave_const[j]
                sumy+=data_vec[j]
            peaks_vec[i] = sumxy/sumy

        # print(peaks_vec)
        return peaks_vec
    
    def adc_filter(self, adc_vec):
        y = medfilt(adc_vec, kernel_size=3)

        # fs = 1000
        # cutoff = 100
        # b,a = butter(N=4,Wn=cutoff/(fs/2),btype='low')
        # y = filtfilt(b,a,y)

        self.filter_visual(adc_vec, y)
        return y
    
    def filter_visual(self, adc_vec, fil_vec):
        x_data = []
        y_data = []
        for i,(fi,adc) in enumerate(zip(fil_vec, adc_vec)):
            if abs(fi-adc)<0.1:
                continue
            x_data.append(self.wave_const[i])
            y_data.append(adc)

            scatter = pg.ScatterPlotItem(x_data, y_data, pen=pg.mkPen(255,255,0), symbol='o', symbolBrush=pg.mkBrush(255,255,0), symbolSize=10)
                
            self.impress_points.append(scatter)
            self.plot1.addItem(scatter)

class ap6150bWindow(QtWidgets.QWidget):
    def __init__(self):
        super().__init__()
        layout = QtWidgets.QGridLayout()
        self.setLayout(layout)

        self.ctrl_panel = QtWidgets.QWidget()
        self.ctrl_layout = QtWidgets.QHBoxLayout()
        self.ctrl_layout.setContentsMargins(5,5,5,5)
        self.ctrl_layout.setSpacing(10)
        self.ctrl_panel.setLayout(self.ctrl_layout)
        self.ctrl_panel.setMaximumHeight(80)

        self.fileText_edit = QLineEdit()
        self.fileText_edit.setPlaceholderText("选择波长数据文件")
        self.fileText_edit.setMinimumWidth(300)
        self.fileText_edit.setMaximumHeight(25)
        self.file_button = QPushButton("...")
        self.file_path = ""

        lab_temperature = QtWidgets.QLabel("温度(℃):")
        self.temperature_text = QtWidgets.QLineEdit("0")
        self.temperature_text.setReadOnly(True)
        self.temperature_text.setMinimumWidth(100)
        self.temperature_text.setMaximumHeight(25)
        self.temperature = 0

        self.com_btn = QtWidgets.QPushButton("开始")
        self.com_btn.setMinimumWidth(100)

        self.clear_btn = QtWidgets.QPushButton("清空日志")
        self.clear_btn.setMinimumWidth(100)

        self.file_button.clicked.connect(self.select_file)
        self.com_btn.clicked.connect(self.ap_thread)
        self.clear_btn.clicked.connect(self.on_clear)

        self.ctrl_layout.addWidget(self.fileText_edit)
        self.ctrl_layout.addWidget(self.file_button)
        self.ctrl_layout.addSpacing(20)
        self.ctrl_layout.addStretch()

        self.ctrl_layout.addWidget(lab_temperature)
        self.ctrl_layout.addWidget(self.temperature_text)
        self.ctrl_layout.addSpacing(20)
        self.ctrl_layout.addStretch()

        self.ctrl_layout.addWidget(self.com_btn)
        self.ctrl_layout.addSpacing(20)
        self.ctrl_layout.addStretch()

        self.ctrl_layout.addWidget(self.clear_btn)

        layout.addWidget(self.ctrl_panel, 0, 0)

        self.printf_area = LogWidget()
        layout.addWidget(self.printf_area)

        self.worker =  APWorker(self.file_path)
        self.worker.log_signal.connect(self.printf_area.log)
        self.worker.temp_signal.connect(self.update_temp)

    def select_file(self):
        self.file_path, _ = QFileDialog.getOpenFileName(
            self,
            "选择Excel文件",
            get_desktop_path(),
            "ALL Files (*);;Text Files (*.xlsx)"
        )
        self.fileText_edit.setText(self.file_path)

    def ap_thread(self):
        global ser
        global ap_open
        global switch_mode_enable

        do_open = False
        do_close = False

        if ap_open == False:
            do_open = True
            ap_open = True
        else: 
            do_close = True
            ap_open = False

        if do_open:
            try:
                ser.open()
            except Exception as e:
                print("Serial Error:",str(e))
                with ap_cond:
                    ap_open = False
                return

            switch_mode_enable = False

            self.worker = APWorker(self.file_path)
            self.worker.log_signal.connect(self.printf_area.log)
            self.worker.temp_signal.connect(self.update_temp)
            self.worker.start()

            self.com_btn.setText("关闭")
        elif do_close:
            self.worker.running = False

            try:
                ser.close()
            except Exception as e:
                print("Serial Error:",str(e))
                with ap_cond:
                    ap_open = True
                return
            
            switch_mode_enable = True

            self.worker.wait()

            self.com_btn.setText("开始")

    def on_clear(self):
        self.printf_area.clear()

    def update_temp(self, _temperature):
        self.temperature = _temperature
        self.temperature_text.setText(f"{_temperature}")

class extraWindow(QtWidgets.QWidget):
    temp_signal = pyqtSignal(float)
    rt_signal = pyqtSignal(int,int)
    error_signal = pyqtSignal(str)

    def __init__(self):
        super().__init__()
        layout = QtWidgets.QVBoxLayout(self)
        layout.setSpacing(15)
        layout.setContentsMargins(20,20,20,20)

        top_layout = QtWidgets.QHBoxLayout()

        head_box = QtWidgets.QGroupBox("通过excel文本导入数据")
        head_layout = QtWidgets.QHBoxLayout()

        self.excel_text = QtWidgets.QLineEdit()
        self.excel_text.setPlaceholderText("Excel复制的文本:(格式:gain\tsoa\tphase\twavelena\twavelenb)...")
        self.excel_text.setMinimumHeight(40)
        self.excel_text.setMinimumWidth(720)

        self.write_btn = QtWidgets.QPushButton("写入")
        self.write_btn.clicked.connect(self.write_data_MCU)
        self.write_btn.setMinimumHeight(40)
        self.write_btn.setMinimumWidth(120)

        head_layout.addWidget(self.excel_text)
        head_layout.addWidget(self.write_btn)

        head_box.setLayout(head_layout)
        head_box.setMaximumHeight(150)

        button_box = QtWidgets.QGroupBox("工作流控制")
        button_layout = QtWidgets.QHBoxLayout()

        self.monitor_btn = QtWidgets.QPushButton("开始")
        self.monitor_btn.clicked.connect(self.on_work)
        self.write_btn.setMinimumHeight(40)
        self.write_btn.setMinimumWidth(120)

        button_layout.addWidget(self.monitor_btn)

        button_box.setLayout(button_layout)
        button_box.setMaximumHeight(150)

        top_layout.addWidget(head_box, 2)
        top_layout.addWidget(button_box, 1)

        layout.addLayout(top_layout)

        body_layout = QtWidgets.QHBoxLayout()

        param_box = QtWidgets.QGroupBox("参数")
        form_left = QtWidgets.QFormLayout()
        form_left.setVerticalSpacing(18)

        def create_param(name):
            edit = QtWidgets.QLineEdit("0")
            edit.setMinimumWidth(180)

            dac_name = QtWidgets.QLabel("实际DAC写入(十进制):")

            dac_text = QtWidgets.QLabel(f"{0}")
            dac_text.setMinimumWidth(80)
            dac_text.setAlignment(QtCore.Qt.AlignRight | QtCore.Qt.AlignVCenter)

            state = QtWidgets.QLabel("● 未写入")
            state.setStyleSheet("color:#d97706")

            h = QtWidgets.QHBoxLayout()
            h.addWidget(edit)
            h.addWidget(dac_name)
            h.addWidget(dac_text)
            h.addWidget(state)

            widget = QtWidgets.QWidget()
            widget.setLayout(h)

            form_left.addRow(name, widget)

            return edit, state, dac_text
        
        (self.gain_text,self.gain_state,self.gain_dac_text),self.gain = create_param("GAIN(mA)"),0
        (self.soa_text,self.soa_state,self.soa_dac_text),self.soa = create_param("SOA(mA)"),0
        (self.phase_text,self.phase_state,self.phase_dac_text),self.phase = create_param("PHASE(mA)"),0
        (self.wavelena_text,self.wavelena_state,self.wavelena_dac_text),self.wavelena = create_param("WAVELENA(mA)"),0
        (self.wavelenb_text,self.wavelenb_state,self.wavelenb_dac_text),self.wavelenb = create_param("WAVELENB(mA)"),0

        param_box.setLayout(form_left)

        monitor_box = QtWidgets.QGroupBox("监测")
        form_right = QtWidgets.QFormLayout()
        form_right.setVerticalSpacing(20)

        def read_only(name):
            e = QtWidgets.QLineEdit(f"{0:.4f}")

            e.setReadOnly(True)
            e.setMinimumWidth(180)

            form_right.addRow(name, e)

            return e

        self.temperature_text,self.temperature = read_only("温度(℃)"),0

        def create_monitor(name):
            e = QtWidgets.QLineEdit("0")
            e.setReadOnly(True)
            e.setMinimumWidth(180)

            v_name = QtWidgets.QLabel("ADC电压:")

            v_text = QtWidgets.QLabel(f"{0:>5}V")

            h = QtWidgets.QHBoxLayout()
            h.addWidget(e)
            h.addWidget(v_name)
            h.addWidget(v_text)

            widget = QtWidgets.QWidget()
            widget.setLayout(h)

            form_right.addRow(name, widget)

            return e, v_text
        
        (self.pdr_text,self.pdr_v),self.pdr = create_monitor("PDR(mA)"),0
        (self.pdt_text,self.pdt_v),self.pdt = create_monitor("PDT(mA)"),0
        self.ratio_rt_text,self.ratio_rt = read_only("PDR/PDT"),0

        monitor_box.setLayout(form_right)

        body_layout.addWidget(param_box,2)
        body_layout.addWidget(monitor_box,1)

        layout.addLayout(body_layout)

        self.excel_text.textChanged.connect(self.excel_adc_value)

        self.gain_text.textChanged.connect(self.gain_transfer)
        self.soa_text.textChanged.connect(self.soa_transfer)
        self.phase_text.textChanged.connect(self.phase_transfer)
        self.wavelena_text.textChanged.connect(self.wavelena_transfer)
        self.wavelenb_text.textChanged.connect(self.wavelenb_transfer)

        self.temp_signal.connect(
            lambda x:self.temperature_text.setText(f"{x:.4f}")
        )
        self.rt_signal.connect(self.rt_transfer)
        self.error_signal.connect(self.show_error)

        self.res_queue = Queue()
        self.exrx_size = 20
        self.running = True
        self.recv_thread = None

    # def showEvent(self, a0):
    #     self.serial_open()
    #     self.start_recv_thread()
    #     return super().showEvent(a0)

    def start_recv_thread(self):
        if self.recv_thread and self.recv_thread.is_alive():
            return
        
        self.running = True
        self.recv_thread = threading.Thread(target=self.serial_recv, daemon=True)
        self.recv_thread.start()

        self.gain_text.textChanged.emit(self.gain_text.text())
        self.soa_text.textChanged.emit(self.soa_text.text())
        self.phase_text.textChanged.emit(self.phase_text.text())
        self.wavelena_text.textChanged.emit(self.wavelena_text.text())
        self.wavelenb_text.textChanged.emit(self.wavelenb_text.text())

    def stop_recv_thread(self):
        self.running = False

        if self.recv_thread:
            self.recv_thread.join(timeout=0.1)
            self.recv_thread = None

    def on_work(self):
        global ser
        global ap_open
        global switch_mode_enable

        do_open = False
        do_close = False

        if ap_open == False:
            do_open = True
            ap_open = True
        else: 
            do_close = True
            ap_open = False

        if do_open:
            try:
                ser.open()
            except Exception as e:
                print("Serial Error:",str(e))
                with ap_cond:
                    ap_open = False
                return

            switch_mode_enable = False

            self.start_recv_thread()

            self.monitor_btn.setText("停止")
        elif do_close:
            self.stop_recv_thread()

            try:
                ser.close()
            except Exception as e:
                print("Serial Error:",str(e))
                with ap_cond:
                    ap_open = True
                return
            
            switch_mode_enable = True

            self.monitor_btn.setText("开始")

    def serial_recv(self):
        self.serial_open()

        while self.running:
            self.serial_open()
            v = ser.read(self.exrx_size)
            if not v:
                continue

            v = list(v)
            if v[0] == 0xFF and v[1] == 0xFF:
                if v[3] == 0x00:
                    self.res_queue.put(v)
                elif v[3] == 0x01:
                    temp_int_high = v[4]
                    temp_int_low = v[5]
                    temp_dec_high = v[6]
                    temp_dec_low = v[7]
                    self.temperature = (temp_int_high<<8) + temp_int_low + \
                                    ((temp_dec_high<<8)+temp_dec_low)*0.0001
                    self.temp_signal.emit(self.temperature)
                elif v[3] == 0x02:
                    _pdt = (v[4]<<8)+v[5]# 先采的ch6是pdt
                    _pdr = (v[6]<<8)+v[7]# 后采的ch7是pdr

                    self.rt_signal.emit(_pdr,_pdt)

        self.serial_close()

    def write_data_MCU(self):
        bWrite = [0] * tx_size
        bWrite[0] = 0xFF
        bWrite[1] = 0xFF
        bWrite[2] = 0x00
        bWrite[3] = dac_type
        bWrite[4] = (self.gain >> 8) & 0xFF
        bWrite[5] = (self.gain >> 0) & 0xFF
        bWrite[6] = (self.soa >> 8) & 0xFF
        bWrite[7] = (self.soa >> 0) & 0xFF
        bWrite[8] = (self.phase >> 8) & 0xFF
        bWrite[9] = (self.phase >> 0) & 0xFF
        bWrite[10] = (self.wavelena >> 8) & 0xFF
        bWrite[11] = (self.wavelena >> 0) & 0xFF
        bWrite[12] = (self.wavelenb >> 8) & 0xFF
        bWrite[13] = (self.wavelenb >> 0) & 0xFF

        cmd = bytes(bWrite)

        serial_write(cmd)

        response = 0
        try:
            response = self.res_queue.get(timeout=1.0)
        except Empty:
            self.error_signal.emit("未写入成功")
            return

        res_gain = (response[4]<<8)+response[5]
        res_soa = (response[6]<<8)+response[7]
        res_phase = (response[8]<<8)+response[9]
        res_wavelena = (response[10]<<8)+response[11]
        res_wavelenb = (response[12]<<8)+response[13]

        res_list, self_list, state_list = [res_gain,res_soa,res_phase,res_wavelena,res_wavelenb],\
                            [self.gain,self.soa,self.phase,self.wavelena,self.wavelenb],\
                            [self.gain_state, self.soa_state,self.phase_state,self.wavelena_state,self.wavelenb_state]

        for i in range(len(res_list)):
            if res_list[i]==self_list[i]:
                state_list[i].setText("● 已写入")
                state_list[i].setStyleSheet("color:#16a34a;")

    def excel_adc_value(self, str):
        cell_list = str.rstrip("\n").split("\t")
        if not len(cell_list) == 5:
            return
        self_text_list =[
            self.gain_text,self.soa_text,self.phase_text,self.wavelena_text,self.wavelenb_text
        ]
        for i in range(len(self_text_list)):
            self_text_list[i].setText(cell_list[i])

    def gain_transfer(self, x):
        x = float(x)

        self.gain = int(
            (1+(x/1000)*18)*100*4096/(2*1.25*160) # U_DAC公式
        ) \
        if dac_type==0 else \
        int(
            x/150*65536 # I_DAC公式
        )

        self.gain_dac_text.setText(f"{self.gain}")
        self.reset_write_state()
    
    def soa_transfer(self, x):
        x = float(x)

        self.soa = \
        int(
            (1+(x/1000)*18)*100*4096/(2*1.25*160) # U_DAC公式
        ) \
        if dac_type==0 else \
        int(
            x/150*65536 # I_DAC公式
        )

        self.soa_dac_text.setText(f"{self.soa}")
        self.reset_write_state()
    
    def phase_transfer(self, x):
        x = float(x)

        self.phase = \
        int(
            (1+(x/1000)*365)*100*4096/(2*1.25*365) # U_DAC公式
        ) \
        if dac_type==0 else \
        int(
            x/20*65536 # I_DAC公式
        )

        self.phase_dac_text.setText(f"{self.phase}")
        self.reset_write_state()
    
    def wavelena_transfer(self, x):
        x = float(x)

        self.wavelena = \
        int(
            (x/1000)*100*50*4096/(2*1.25*53.6) # U_DAC公式
        ) \
        if dac_type==0 else \
        int(
            x/80*65536 # I_DAC公式
        )

        self.wavelena_dac_text.setText(f"{self.wavelena}")
        self.reset_write_state()
    
    def wavelenb_transfer(self, x):
        x = float(x)

        self.wavelenb = \
        int(
            (x/1000)*100*169*4096/(2*1.25*475) # U_DAC公式
        ) \
        if dac_type==0 else \
        int(
            x/80*65536 # I_DAC公式
        )

        self.wavelenb_dac_text.setText(f"{self.wavelenb}")
        self.reset_write_state()

    def rt_transfer(self, r, t):
        v7 = r*2.5/4095
        v6 = t*2.5/4095
        self.pdr = (1.25-v7)/2
        self.pdt = (1.25-v6)/2
        self.ratio_rt = self.pdr/self.pdt

        self.pdr_v.setText(f"{v7:.2f}V")
        self.pdt_v.setText(f"{v6:.2f}V")
        self.pdr_text.setText(f"{self.pdr:.4f}")
        self.pdt_text.setText(f"{self.pdt:.4f}")
        self.ratio_rt_text.setText(f"{self.ratio_rt:.3f}")

    def reset_write_state(self):
        states = [self.gain_state, self.soa_state,self.phase_state,self.wavelena_state,self.wavelenb_state]
        for s in states:
            s.setText("● 未写入")
            s.setStyleSheet("color:#d97706")

    def show_error(self, str):
        QMessageBox.critical(self, "crash", f"异常:\n{str}")

    def serial_open(self):
        try:
            if not ser.is_open:
                ser.open()
            ser_open = True
        except Exception as e:
            self.error_signal.emit(str(e))
            return
        
    def serial_close(self):
        try:
            ser.close()
            ser_open = False
        except Exception as e:
            self.error_signal.emit(str(e))

if __name__=="__main__":
    QtWidgets.QApplication.setAttribute(QtCore.Qt.AA_EnableHighDpiScaling)
    QtWidgets.QApplication.setAttribute(QtCore.Qt.AA_UseHighDpiPixmaps)
    app = QtWidgets.QApplication(sys.argv)
    app.setStyleSheet("""
        QPushButton, QLabel {
            font-size: 12pt;
            font-family: Microsoft YaHei;
        }
    """)
    win = MainWindow()
    win.show()
    sys.exit(app.exec_())


