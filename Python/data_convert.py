import os
import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
from openpyxl import load_workbook
import yaml

excel_path = ""

def to_int_str(v):
    """把单元格值转成“截断后的整数”字符串：不四舍五入，仅保留整数部分。"""
    if v is None:
        return ""

    # openpyxl 可能读到 int/float/str
    # 1) 数字类型：直接向0截断
    if isinstance(v, (int, float)):
        return str(int(v))

    # 2) 字符串：尽量转成数字后再截断，否则原样去空格
    s = str(v).strip()
    if s == "":
        return ""
    try:
        return str(int(float(s)))  # 先float再int=向0截断
    except ValueError:
        return s  # 不是数字就原样输出
    
def split_number(x):
    try:
        s = f"{float(x):.3f}"
        integer_part, decimal_part = s.split('.')
        return int(integer_part), int(decimal_part)
    except Exception as e:
        print(f"excel读取的值:{x}")
        print(e)

def dac_main(ws):
    lines = []
    for row in range(2, ws.max_row + 1):
        # 从1开始计数
        g = ws.cell(row=row, column=7).value
        h = ws.cell(row=row, column=8).value
        i = ws.cell(row=row, column=9).value
        j = ws.cell(row=row, column=10).value
        k = ws.cell(row=row, column=11).value

        if g is None and h is None and i is None:
            continue

        g_s, h_s, i_s, j_s, k_s = to_int_str(g), to_int_str(h), to_int_str(i), to_int_str(j), to_int_str(k)
        lines.append(f"{{{g_s}, {h_s}, {i_s}, {j_s}, {k_s}}},")

    default_name = os.path.splitext(os.path.basename(excel_path))[0] + "_dac"
    save_path = filedialog.asksaveasfilename(
        title="保存为TXT",
        defaultextension=".txt",
        initialfile=default_name,
        filetypes=[("Text files", "*.txt"), ("All files", "*.*")],
    )
    if not save_path:
        return

    try:
        with open(save_path, "w", encoding="utf-8") as f:
            f.write("\n".join(lines))
        messagebox.showinfo("完成", f"已生成TXT：\n{save_path}")
    except Exception as ex:
        messagebox.showerror("保存失败", f"保存TXT时出错：\n{ex}")

def wave_main(ws):
    results = []
    yaml_results = []
    for row in range(2, ws.max_row + 1):
        # 空值跳过（你也可以改成写 {0,0}）
        value = ws.cell(row=row, column=13).value
        if pd.isna(value):
            continue
        integer_part, decimal_part = split_number(value)
        results.append(f"{{{integer_part},{decimal_part}}},")
        yaml_results.append([integer_part,decimal_part])

    default_name = os.path.splitext(os.path.basename(excel_path))[0] + "_wave"
    save_path = filedialog.asksaveasfilename(
        title="保存为txt",
        defaultextension=".txt",
        initialfile=default_name,
        filetypes=[("Text files", "*.txt")]
    )

    # print(results)
    if save_path:
        try:
            with open(save_path, "w", encoding="utf-8") as f:
                for line in results:
                    f.write(line + "\n")
            messagebox.showinfo("完成", f"已生成TXT：\n{save_path}")
        except Exception as ex:
            messagebox.showerror("保存失败", f"保存TXT时出错：\n{ex}")

    yaml_path = filedialog.asksaveasfilename(
        title="保存为yaml",
        defaultextension=".yaml",
        initialfile=default_name,
        filetypes=[("Yaml files", "*.yaml")]
    )

    if yaml_path:
        yaml_data = {"Wave_DATA":yaml_results}
        try:
            with open(yaml_path, "w", encoding="utf-8") as f:
                yaml.safe_dump(
                    yaml_data,
                    f,
                    allow_unicode=True,
                    sort_keys=False,
                    default_flow_style=False
                )
            messagebox.showinfo("完成", f"已生成TXT：\n{yaml_path}")
        except Exception as ex:
            messagebox.showerror("保存失败", f"保存TXT时出错：\n{ex}")

def excel_main():
    root = tk.Tk()
    root.withdraw()
    root.update()

    global excel_path
    excel_path = filedialog.askopenfilename(
        title="选择Excel文件",
        filetypes=[
            ("Excel files", "*.xlsx *.xlsm *.xltx *.xltm"),
            ("All files", "*.*"),
        ],
    )
    if not excel_path:
        return

    try:
        wb = load_workbook(excel_path, data_only=True)
        ws = wb.active
    except Exception as ex:
        messagebox.showerror("读取失败", f"读取Excel时出错：\n{ex}")
        return
    
    dac_main(ws)
    wave_main(ws)

if __name__ == "__main__":
    excel_main()