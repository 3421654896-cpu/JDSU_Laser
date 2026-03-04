# -*- coding: utf-8 -*-
import os
import sys
import time
import math
from decimal import Decimal, ROUND_FLOOR
from typing import Optional, Tuple, Dict, Any
from copy import copy as shallow_copy

import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog

from openpyxl import load_workbook, Workbook
from openpyxl.utils import column_index_from_string


def log(msg: str):
    print(msg)
    sys.stdout.flush()


def get_desktop_path():
    home = os.path.expanduser("~")
    for name in ("Desktop", "桌面"):
        p = os.path.join(home, name)
        if os.path.isdir(p):
            return p
    return home


def to_decimal(v) -> Optional[Decimal]:
    """
    尽量把单元格值转成 Decimal。
    遇到 NaN/Infinity 等“非有限数”一律返回 None，
    防止后续比较触发 decimal.InvalidOperation
    """
    if v is None:
        return None

    if isinstance(v, (int, float)):
        if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
            return None
        try:
            d = Decimal(str(v))
        except Exception:
            return None
        if (not d.is_finite()) or d.is_nan():
            return None
        return d

    if isinstance(v, str):
        s = v.strip()
        if not s:
            return None
        sl = s.lower()
        if sl in ("nan", "+nan", "-nan", "inf", "+inf", "-inf", "infinity", "+infinity", "-infinity"):
            return None
        try:
            d = Decimal(s)
        except Exception:
            return None
        if (not d.is_finite()) or d.is_nan():
            return None
        return d

    return None


def mw_to_dbm(mw: Decimal) -> Optional[Decimal]:
    """mW -> dBm；mw<=0 返回 None"""
    if mw is None:
        return None
    if mw <= 0:
        return None
    return Decimal(str(10.0 * math.log10(float(mw))))


def should_skip_by_abs_dbm_delta_lt_threshold(mw1, mw3, threshold_db: Decimal) -> bool:
    """
    True => 跳过（不参与筛选）：
      - abs(dBm(mw1) - dBm(mw3)) < threshold_db

    说明：
      - 任意一列无法转成数值、或 <=0（无法转dBm） => 不跳过（保留参与筛选）
      - mw1/mw3 单位按 mW
    """
    m1 = to_decimal(mw1)
    m3 = to_decimal(mw3)
    if m1 is None or m3 is None:
        return False

    d1 = mw_to_dbm(m1)
    d3 = mw_to_dbm(m3)
    if d1 is None or d3 is None:
        return False

    delta = abs(d1 - d3)
    return delta < threshold_db


def nearest_even_cent_key_and_dist(x: Decimal) -> Tuple[int, Decimal]:
    """key=最近偶数分(单位=分)，dist=|x*100-key|"""
    m = x * Decimal("100")
    floor_int = int(m.to_integral_value(rounding=ROUND_FLOOR))

    n_down = floor_int if (floor_int % 2 == 0) else (floor_int - 1)
    n_up = n_down + 2

    dist_down = abs(m - Decimal(n_down))
    dist_up = abs(m - Decimal(n_up))

    n_best = n_down if dist_down <= dist_up else n_up
    return n_best, abs(m - Decimal(n_best))


def select_excel_file(root) -> str:
    return filedialog.askopenfilename(
        title="请选择要筛选的Excel文件",
        filetypes=[("Excel文件", "*.xlsx *.xlsm *.xltx *.xltm"), ("所有文件", "*.*")],
        initialdir=get_desktop_path()
    )


def select_save_path(root, default_name: str) -> str:
    return filedialog.asksaveasfilename(
        title="请选择筛选后Excel的保存位置",
        defaultextension=".xlsx",
        initialdir=get_desktop_path(),
        initialfile=default_name,
        filetypes=[("Excel文件", "*.xlsx")]
    )


def _fmt_eta(seconds: float) -> str:
    if seconds < 0 or seconds == float("inf"):
        return "?"
    m, s = divmod(int(seconds), 60)
    h, m = divmod(m, 60)
    if h > 0:
        return f"{h}h{m:02d}m"
    if m > 0:
        return f"{m}m{s:02d}s"
    return f"{s}s"


def pick_best_rows_by_target(
    in_path: str,
    sheet: Optional[str],
    header: bool,
    col_idx: int,
    progress_every_rows: int = 2000,
    diff_threshold_db: Decimal = Decimal("15"),
    diag_range_nm: Tuple[Decimal, Decimal] = (Decimal("1525"), Decimal("1565")),
) -> Tuple[Dict[int, int], Dict[str, Any]]:
    """
    扫描一遍：
      1) 先统计“每个目标点key”有哪些行落入（用于缺失原因归类）
      2) 再跳过：abs(dBm(mW1)-dBm(mW3)) < diff_threshold_db 的行
      3) 剩下参与筛选：按 dist/光强 规则选最优行

    返回：
      - best_map: {key: best_row_index}
      - diag: 用于缺失原因归类的统计
    """
    wb_val = load_workbook(in_path, read_only=True, data_only=True)
    ws_val = wb_val[sheet] if sheet else wb_val.active

    start_row = 2 if header else 1
    max_row = ws_val.max_row

    best: Dict[int, Tuple[Decimal, Decimal, int]] = {}

    t0 = time.time()
    scanned = 0
    skipped = 0

    intensity_col_idx = col_idx + 1  # 光强：波长后第1列（用于 dist 相同决策）
    diff_col1_idx = col_idx + 1      # mW1：波长后第1列
    diff_col3_idx = col_idx + 3      # mW3：波长后第3列

    # ====== 诊断统计：只统计指定范围内的key（1525-1565） ======
    range_start_nm, range_end_nm = diag_range_nm
    start_key = int((range_start_nm * Decimal("100")).to_integral_value(rounding=ROUND_FLOOR))
    end_key = int((range_end_nm * Decimal("100")).to_integral_value(rounding=ROUND_FLOOR))
    if start_key % 2 != 0:
        start_key += 1
    if end_key % 2 != 0:
        end_key -= 1

    total_seen: Dict[int, int] = {}   # 落入该key的总行数（无论是否跳过）
    kept_seen: Dict[int, int] = {}    # 落入该key且参与筛选的行数
    skipped_seen: Dict[int, int] = {} # 落入该key但被跳过的行数
    invalid_wavelength_rows = 0       # 波长列无法解析的行数（无法归到任何key）
    # ===========================================================

    for row_idx, row_vals in enumerate(ws_val.iter_rows(min_row=start_row, values_only=True), start=start_row):
        scanned += 1

        # 先取波长，算 key（用于统计缺失原因）
        wl = to_decimal(row_vals[col_idx - 1] if (col_idx - 1) < len(row_vals) else None)
        if wl is None:
            invalid_wavelength_rows += 1
            # 进度打印
            if scanned % progress_every_rows == 0 or row_idx == max_row:
                now = time.time()
                elapsed = now - t0
                speed = scanned / elapsed if elapsed > 0 else 0.0
                remaining_rows = (max_row - row_idx)
                eta = remaining_rows / speed if speed > 0 else float("inf")
                log(
                    f"[扫描] 行 {row_idx}/{max_row} ({row_idx/max_row*100:.1f}%) | "
                    f"目标点数={len(best)} | 跳过行={skipped} | "
                    f"波长无效行={invalid_wavelength_rows} | {speed:.0f} 行/s | ETA {_fmt_eta(eta)}"
                )
            continue

        key, dist = nearest_even_cent_key_and_dist(wl)

        # 只对 1525-1565 范围内的key做归因统计
        if start_key <= key <= end_key:
            total_seen[key] = total_seen.get(key, 0) + 1

        # 再取跳过判定列（mW1/mW3）
        mw1 = row_vals[diff_col1_idx - 1] if (diff_col1_idx - 1) < len(row_vals) else None
        mw3 = row_vals[diff_col3_idx - 1] if (diff_col3_idx - 1) < len(row_vals) else None

        if should_skip_by_abs_dbm_delta_lt_threshold(mw1, mw3, diff_threshold_db):
            skipped += 1
            if start_key <= key <= end_key:
                skipped_seen[key] = skipped_seen.get(key, 0) + 1
            if scanned % progress_every_rows == 0 or row_idx == max_row:
                now = time.time()
                elapsed = now - t0
                speed = scanned / elapsed if elapsed > 0 else 0.0
                remaining_rows = (max_row - row_idx)
                eta = remaining_rows / speed if speed > 0 else float("inf")
                log(
                    f"[扫描] 行 {row_idx}/{max_row} ({row_idx/max_row*100:.1f}%) | "
                    f"目标点数={len(best)} | 跳过行={skipped} | "
                    f"波长无效行={invalid_wavelength_rows} | {speed:.0f} 行/s | ETA {_fmt_eta(eta)}"
                )
            continue

        # 参与筛选计数
        if start_key <= key <= end_key:
            kept_seen[key] = kept_seen.get(key, 0) + 1

        # 光强列（dist 相同用它做决策）
        intensity = to_decimal(row_vals[intensity_col_idx - 1] if (intensity_col_idx - 1) < len(row_vals) else None)
        intensity = intensity if intensity is not None else Decimal("-Infinity")

        if key not in best:
            best[key] = (dist, intensity, row_idx)
        else:
            best_dist, best_intensity, best_row = best[key]
            if dist < best_dist:
                best[key] = (dist, intensity, row_idx)
            elif dist == best_dist and intensity > best_intensity:
                best[key] = (dist, intensity, row_idx)

        if scanned % progress_every_rows == 0 or row_idx == max_row:
            now = time.time()
            elapsed = now - t0
            speed = scanned / elapsed if elapsed > 0 else 0.0
            remaining_rows = (max_row - row_idx)
            eta = remaining_rows / speed if speed > 0 else float("inf")
            log(
                f"[扫描] 行 {row_idx}/{max_row} ({row_idx/max_row*100:.1f}%) | "
                f"目标点数={len(best)} | 跳过行={skipped} | "
                f"波长无效行={invalid_wavelength_rows} | {speed:.0f} 行/s | ETA {_fmt_eta(eta)}"
            )

    if not best:
        raise RuntimeError("筛选后未找到可解析为数字的值（可能都被跳过或为空），无法筛选。")

    best_map = {k: r for k, (d, inten, r) in best.items()}

    diag = {
        "range_nm": (range_start_nm, range_end_nm),
        "range_key": (start_key, end_key),
        "total_seen": total_seen,
        "kept_seen": kept_seen,
        "skipped_seen": skipped_seen,
        "invalid_wavelength_rows": invalid_wavelength_rows,
        "skipped_rows_total": skipped,
        "scanned_rows_total": scanned,
    }
    return best_map, diag


def copy_sheet_settings(ws_in, ws_out):
    ws_out.freeze_panes = ws_in.freeze_panes

    for col_letter, dim in ws_in.column_dimensions.items():
        out_dim = ws_out.column_dimensions[col_letter]
        out_dim.width = dim.width
        out_dim.hidden = dim.hidden
        out_dim.outlineLevel = dim.outlineLevel
        out_dim.bestFit = dim.bestFit

    try:
        ws_out.sheet_format.defaultRowHeight = ws_in.sheet_format.defaultRowHeight
    except Exception:
        pass


def copy_cell_style_and_value(in_style_cell, in_value_cell, out_cell):
    out_cell.value = in_value_cell.value

    out_cell.font = shallow_copy(in_style_cell.font)
    out_cell.fill = shallow_copy(in_style_cell.fill)
    out_cell.border = shallow_copy(in_style_cell.border)
    out_cell.alignment = shallow_copy(in_style_cell.alignment)
    out_cell.number_format = in_style_cell.number_format
    out_cell.protection = shallow_copy(in_style_cell.protection)

    if in_style_cell.hyperlink is not None:
        out_cell.hyperlink = shallow_copy(in_style_cell.hyperlink)
    if in_style_cell.comment is not None:
        out_cell.comment = shallow_copy(in_style_cell.comment)


def copy_kept_rows_style_valueonly(
    in_path: str,
    out_path: str,
    sheet: Optional[str],
    kept_rows_sorted: list,
    header: bool,
    progress_every_rows: int = 200
):
    wb_style = load_workbook(in_path, data_only=False)
    ws_style = wb_style[sheet] if sheet else wb_style.active

    wb_val = load_workbook(in_path, data_only=True)
    ws_val = wb_val[sheet] if sheet else wb_val.active

    wb_out = Workbook()
    if wb_out.active and wb_out.active.title == "Sheet":
        wb_out.remove(wb_out.active)
    ws_out = wb_out.create_sheet(title=ws_style.title[:31])

    copy_sheet_settings(ws_style, ws_out)

    rows_to_copy = []
    if header:
        rows_to_copy.append(1)
    rows_to_copy.extend(kept_rows_sorted)

    row_map = {old_r: new_r for new_r, old_r in enumerate(rows_to_copy, start=1)}
    total = len(rows_to_copy)

    t0 = time.time()
    for i, old_r in enumerate(rows_to_copy, start=1):
        new_r = row_map[old_r]

        in_rd = ws_style.row_dimensions.get(old_r)
        out_rd = ws_out.row_dimensions[new_r]
        if in_rd is not None:
            out_rd.height = in_rd.height
            out_rd.hidden = in_rd.hidden
            out_rd.outlineLevel = in_rd.outlineLevel

        for c in range(1, ws_style.max_column + 1):
            in_style_cell = ws_style.cell(row=old_r, column=c)
            in_value_cell = ws_val.cell(row=old_r, column=c)
            out_cell = ws_out.cell(row=new_r, column=c)
            copy_cell_style_and_value(in_style_cell, in_value_cell, out_cell)

        if i % progress_every_rows == 0 or i == total:
            elapsed = time.time() - t0
            speed = i / elapsed if elapsed > 0 else 0.0
            eta = (total - i) / speed if speed > 0 else float("inf")
            log(f"[复制] {i}/{total} ({i/total*100:.1f}%) | {speed:.0f} 行/s | ETA {_fmt_eta(eta)}")

    kept_set = set(rows_to_copy)
    for merged in list(ws_style.merged_cells.ranges):
        min_row = merged.min_row
        max_row = merged.max_row

        if any(rr not in kept_set for rr in range(min_row, max_row + 1)):
            continue

        new_min = row_map[min_row]
        new_max = row_map[max_row]
        if (new_max - new_min) != (max_row - min_row):
            continue

        ws_out.merge_cells(
            start_row=new_min, end_row=new_max,
            start_column=merged.min_col, end_column=merged.max_col
        )

    wb_out.save(out_path)


def _parse_decimal_or_default(s: str, default: Decimal) -> Decimal:
    if s is None:
        return default
    ss = str(s).strip()
    if ss == "":
        return default
    try:
        return Decimal(ss)
    except Exception:
        return default


def _print_missing_grouped(best_map: Dict[int, int], diag: Dict[str, Any]):
    """按原因归类打印缺失目标点"""
    (start_nm, end_nm) = diag["range_nm"]
    (start_key, end_key) = diag["range_key"]
    total_seen = diag["total_seen"]
    kept_seen = diag["kept_seen"]
    skipped_seen = diag["skipped_seen"]

    missing_no_data = []   # A类：没有任何行落入
    missing_all_skipped = []  # B类：有行落入但全被跳过（kept=0）

    for key in range(start_key, end_key + 1, 2):
        if key in best_map:
            continue

        t = total_seen.get(key, 0)
        k = kept_seen.get(key, 0)
        s = skipped_seen.get(key, 0)

        wl = Decimal(key) / Decimal("100")
        if t == 0:
            missing_no_data.append(wl)
        else:
            # 有落入，但没产出best，基本就是全被跳过（或极端情况下全部强度/波长问题，这里也归到此类）
            missing_all_skipped.append((wl, t, s, k))

    log(f"==== 目标点缺失归因（{start_nm}-{end_nm}，步长0.02nm）====")
    log(f"波长列无法解析的行数（无法归到任何目标点）：{diag['invalid_wavelength_rows']}")
    log(f"总扫描行：{diag['scanned_rows_total']} | 总跳过行（阈值条件）：{diag['skipped_rows_total']}")
    log("--------------------------------------------------")

    log(f"A类【无数据落入该目标点】数量：{len(missing_no_data)}")
    if missing_no_data:
        line = []
        for i, wl in enumerate(missing_no_data, 1):
            line.append(f"{wl:.2f}")
            if i % 20 == 0:
                log("A类缺失: " + ", ".join(line))
                line = []
        if line:
            log("A类缺失: " + ", ".join(line))

    log("--------------------------------------------------")

    log(f"B类【有数据落入但全部被跳过】数量：{len(missing_all_skipped)}")
    if missing_all_skipped:
        # B类建议打印更详细：每个点落入多少行、跳过多少行
        # 仍然按行数避免刷屏：每行10个点
        buf = []
        for i, (wl, t, s, k) in enumerate(missing_all_skipped, 1):
            buf.append(f"{wl:.2f}(落入{t},跳过{s},保留{k})")
            if i % 10 == 0:
                log("B类缺失: " + " | ".join(buf))
                buf = []
        if buf:
            log("B类缺失: " + " | ".join(buf))

    log("==== 缺失归因打印结束 ====")


def main():
    root = tk.Tk()
    root.withdraw()
    try:
        root.attributes("-topmost", True)
    except Exception:
        pass

    in_path = select_excel_file(root)
    if not in_path:
        root.destroy()
        return

    ext = os.path.splitext(in_path)[1].lower()
    if ext == ".xls":
        messagebox.showerror("格式不支持", "openpyxl 不支持 .xls，请先另存为 .xlsx 或 .xlsm。")
        root.destroy()
        return

    base = os.path.splitext(os.path.basename(in_path))[0]
    out_path = select_save_path(root, f"{base}_filtered.xlsx")
    if not out_path:
        root.destroy()
        return

    header = messagebox.askyesno("表头设置", "第1行是否是表头？\n（是：保留第1行）")

    sheet = simpledialog.askstring("工作表名称", "请输入工作表名称（留空=使用当前激活表）：")
    sheet = (sheet.strip() if sheet else "")
    sheet = sheet if sheet else None

    col_letter = simpledialog.askstring("列选择", "请输入要处理的列字母（默认 I；例如 L）：")
    col_letter = (col_letter.strip().upper() if col_letter else "I")
    try:
        col_idx = column_index_from_string(col_letter)
    except Exception:
        messagebox.showerror("列输入错误", f"列字母无效：{col_letter}")
        root.destroy()
        return

    # 阈值输入框（单位：dB，等价于 dBm 差值的绝对值）
    thr_str = simpledialog.askstring(
        "跳过阈值(dB)",
        "请输入跳过阈值(dB)：\n当 abs(dBm(mW1) - dBm(mW3)) < 阈值 时，该行跳过不参与筛选。\n\n留空默认 15",
    )
    diff_threshold_db = _parse_decimal_or_default(thr_str, Decimal("15"))

    scan_every = 2000
    copy_every = 200

    try:
        log(f"==== 开始扫描：跳过 abs(dBm(mW1)-dBm(mW3)) < {diff_threshold_db} dB 的行；其它逻辑不变 ====")
        best_map, diag = pick_best_rows_by_target(
            in_path, sheet, header, col_idx,
            progress_every_rows=scan_every,
            diff_threshold_db=diff_threshold_db,
            diag_range_nm=(Decimal("1525"), Decimal("1565")),
        )

        # ====== 新增：按原因归类打印缺失目标点 ======
        _print_missing_grouped(best_map, diag)

        kept_rows = sorted(best_map.values())
        log(f"==== 扫描完成：目标点数量={len(best_map)}，保留行数={len(kept_rows)}（不含表头） ====")

        log("==== 开始复制：保留样式，但只写值（不复制公式） ====")
        copy_kept_rows_style_valueonly(in_path, out_path, sheet, kept_rows, header, progress_every_rows=copy_every)

        messagebox.showinfo(
            "完成",
            f"已输出：\n{out_path}\n\n"
            f"目标点数量：{len(best_map)}\n"
            f"保留数据行数：{len(kept_rows)}（不含表头）\n"
            f"处理列（波长列）：{col_letter}\n"
            f"光强列：{col_letter} 后第1列\n"
            f"跳过条件：abs(dBm(mW1)-dBm(mW3)) < {diff_threshold_db} dB\n"
            f"新文件：样式保留、只复制值（不复制公式）"
        )
        log("完成，输出：" + out_path)

    except Exception as e:
        messagebox.showerror("失败", str(e))
        raise
    finally:
        root.destroy()


if __name__ == "__main__":
    main()
