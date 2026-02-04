# -*- coding: utf-8 -*-
"""
AQ6150B 严格等间隔采样（BUS 触发 + *OPC? 确认），每 0.85 s 采集峰值波长与光功率
ADDR = "GPIB0::7::INSTR"
依赖：pip install pyvisa pyvisa-py pyserial  （或使用 NI-VISA 后端）
"""

import os
import csv
import time
from datetime import datetime
import pyvisa
import serial
import time
import math
from openpyxl import load_workbook, Workbook
import tkinter as tk
from tkinter import filedialog

data_send = 0x11
data_get = 0x21
received_data = 0x00
ser = serial.Serial('COM6', 2000000, timeout=1)

ADDR = "GPIB0::7::INSTR"
SAMPLE_PERIOD_S = 2
CSV_BASENAME = f"AQ6150B_log_{time.strftime('%H%M%S')}.csv"

def get_desktop_path():
    home = os.path.expanduser("~")
    for name in ("Desktop", "桌面"):
        p = os.path.join(home, name)
        if os.path.isdir(p):
            return p
    return home

def select_excel_file():
    """弹窗选择Excel文件"""
    root = tk.Tk()
    root.withdraw()  # 隐藏主窗口
    
    file_path = filedialog.askopenfilename(
        title="请选择Excel文件",
        filetypes=[
            ("Excel文件", "*.xlsx *.xls"),
            ("所有文件", "*.*")
        ],
        initialdir=get_desktop_path()  # 默认打开桌面
    )
    
    root.destroy()
    return file_path

def try_write(inst, cmd):
    try:
        inst.write(cmd)
        return True
    except Exception:
        return False

def serial_write(info=data_send):
    # 写入数据（需要编码为bytes类型）
    # print(len(info))
    ser.write(info)
    # print(f"发送数据: {info}")

def serial_read():
    global received_data
    received_data = int.from_bytes(ser.read(1),'big')
    if received_data:
        print(f"接收到数据: {received_data}")
    # else:
    #     print("未接收到数据")
    return received_data

def excel_operate(iter_excel):
    try:
        row = next(iter_excel)

        readPhase = int(row[4])
        readwaveA = int(row[5])
        readwaveB = int(row[6])
        print("---excel----")
        print(readPhase)
        print(readwaveA)
        print(readwaveB)
        print("---excel----")

        bWrite = [0]*8

        bWrite[0] = 0xFF
        bWrite[1] = 0xFF
        bWrite[2] = (readPhase>>8)&0xFF
        bWrite[3] = (readPhase>>0)&0xFF
        bWrite[4] = (readwaveA>>8)&0xFF
        bWrite[5] = (readwaveA>>0)&0xFF
        bWrite[6] = (readwaveB>>8)&0xFF
        bWrite[7] = (readwaveB>>0)&0xFF
        
        return bytes(bWrite)
    except Exception:
        return None
    

def main():
    # 弹窗选择Excel文件
    excel_path = select_excel_file()
    
    if not excel_path:
        print("未选择文件，程序退出")
        return
    
    print(f"已选择文件: {excel_path}")
    
    try:
        wb_in = load_workbook(excel_path, read_only=True, data_only=True)
        ws_in = wb_in.active
        iter_excel = ws_in.iter_rows(min_row=2, values_only=True)
    except Exception as e:
        print(f"打开Excel文件失败: {e}")
        return

    rm = pyvisa.ResourceManager()
    inst = rm.open_resource(ADDR)
    inst.timeout = 10000  # ms
    inst.read_termination = "\n"
    inst.write_termination = "\n"

    # ----- 基本通信确认 -----
    try:
        idn = inst.query("*IDN?")
        print("Connected:", idn.strip())
    except Exception:
        print("提示：*IDN? 不响应也没关系。")

    # ----- 进入触发采样模式（严格等间隔）-----
    try_write(inst, ":INIT:CONT OFF")
    try_write(inst, ":TRIG:SOUR BUS")
    try_write(inst, ":FORM:NDAT 0NM")
    try_write(inst, ":UNIT:POW DBM")

    # ----- 选择查询命令（优先 FETC 的波长/功率）-----
    query_variants = [
        (":FETC:SCAL:POW:WAV?", ":FETC:SCAL:POW?"),
        (":FETC:WAV?", ":FETC:POW?"),
    ]
    chosen = None
    for wav_cmd, pow_cmd in query_variants:
        try:
            inst.write(":INIT")
            inst.write("*TRG")
            _ = inst.query("*OPC?")
            _ = inst.query(wav_cmd)
            _ = inst.query(pow_cmd)
            chosen = (wav_cmd, pow_cmd)
            break
        except Exception:
            continue

    if not chosen:
        print("错误：没有可用的 FETCh 查询组合。")
        try:
            inst.close()
        finally:
            rm.close()
        return

    wav_cmd, pow_cmd = chosen
    print(f"使用指令组合：{wav_cmd} + {pow_cmd}")

    # ----- 发送手动命令 -----
    debugCommand = [0xFE,0x55,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0x00,0xEF]
    serial_write(bytes(debugCommand))

    # ----- 打开 CSV 并开始严格等间隔采样 -----
    desktop = get_desktop_path()
    csv_path = os.path.join(desktop, CSV_BASENAME)
    file_exists = os.path.isfile(csv_path)

    count=0
    with open(csv_path, "a", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        if not file_exists:
            writer.writerow(["timestamp_iso", "wavelength_nm", "power_dBm"])

        print(f"开始记录（严格等间隔触发），每 {SAMPLE_PERIOD_S:.3f} s 采样一次。按 Ctrl+C 结束。")

        next_t = time.perf_counter()

        try:
            while True:
                regDataCommand = excel_operate(iter_excel)
                if regDataCommand==None: 
                    break
                serial_write(regDataCommand)
                # read = ser.read(8)
                # print(list(read))
                # time.sleep(1)
                # continue

                while serial_read()!=int(0x21):
                    print("返回错误:",received_data)
                    continue
                    # print(received_data)
                # next_t += SAMPLE_PERIOD_S
                # time.sleep(max(0.0, next_t - time.perf_counter()))
                
                count+=1
                print("计数:"+str(count))

                ts = datetime.now().isoformat(timespec="milliseconds")
                try:
                    inst.write(":INIT")
                    inst.write("*TRG")
                    _ = inst.query("*OPC?")
                    wav_m_str = inst.query(wav_cmd)
                    pow_str = inst.query(pow_cmd)
                except Exception:
                    wav_m_str = "nan"
                    pow_str = "nan"

                try:
                    wav_nm = float(wav_m_str) * 1e9
                except Exception:
                    wav_nm = float("nan")

                try:
                    power_dbm = float(pow_str)
                except Exception:
                    power_dbm = float("nan")

                
                # #writer.writerow([ts, f"{wav_nm:.6f}", f"{power_dbm:.6f}"])
                # writer.writerow([ts, f"{wav_nm:.3f}", f"{power_dbm:.3f}"])
                # f.flush()
                
                # 截断函数
                def trunc3(x):
                    if math.isnan(x):
                        return "nan"
                    return f"{math.trunc(x * 1000) / 1000.0:.3f}"

                # 写入 CSV
                writer.writerow([ts, trunc3(wav_nm), trunc3(power_dbm)])
                f.flush()

                time.sleep(0.001)

        except KeyboardInterrupt:
            stopCommand = bytes([0xFE,0x07,0x00,0x00,0x00,0x00,0x00,0x00,
                                       0x00,0x00,0x00,0x00,0x00,0x00,0xEF])
            print(f"\n已停止。数据已保存至：{csv_path}")

    try:
        inst.close()
    finally:
        rm.close()
        print(f"\n已结束。数据已保存至：{csv_path}")

if __name__ == "__main__":
    main()
