# -*- coding: utf-8 -*-
"""
AQ6150B：按串口 ACK 推进（不严格等间隔）
功能：
- 打印 Excel 读取到的数据（可控频率）
- 计数:xxx（还原）
- 读取两强峰波长(nm)与功率（写入 mW）

修改点（按你的要求）：
- ACK 不成功就一直重发同一行，不推进 Excel（不会丢行）
"""

import os
import csv
import time
import math
from datetime import datetime

import pyvisa
import serial
import openpyxl
from openpyxl import load_workbook
import tkinter as tk
from tkinter import filedialog

# ====== 参数（按需改）======
ADDR = "GPIB0::7::INSTR"
COM_PORT = "COM6"
BAUD = 2000000

PRINT_EXCEL_EVERY = 1      # 每隔N行打印一次Excel数据（1=每行都打印）
PRINT_ACK = True           # 打印接收到的ACK

VISA_TIMEOUT_MS = 3000
VISA_RETRY = 2

FLUSH_EVERY_N = 10
XLSX_BASENAME = f"AQ6150B_log_{time.strftime('%H%M%S')}.xlsx"

ACK_VALUE = 0x21
ACK_RESEND_SLEEP_S = 0.001  # 每次重发后短暂停一下，避免占满CPU
# ==========================

ser = serial.Serial(COM_PORT, BAUD, timeout=1)

def get_desktop_path():
    home = os.path.expanduser("~")
    for name in ("Desktop", "桌面"):
        p = os.path.join(home, name)
        if os.path.isdir(p):
            return p
    return home

def select_excel_file():
    root = tk.Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename(
        title="请选择Excel文件",
        filetypes=[("Excel文件", "*.xlsx *.xls"), ("所有文件", "*.*")],
        initialdir=get_desktop_path()
    )
    root.destroy()
    return file_path

def serial_write(info: bytes):
    ser.write(info)

def wait_ack_0x21_forever(regDataCommand: bytes) -> None:
    """
    一直等待直到收到 ACK=0x21。
    期间如果没收到或收到的不是0x21，就重发同一条命令，不推进。
    """
    while True:
        b = ser.read(1)
        if b:
            v = int.from_bytes(b, "big")
            if PRINT_ACK:
                print(f"接收到数据: {v}")
            if v == ACK_VALUE:
                return
        # 没收到或不是0x21 -> 重发
        serial_write(regDataCommand)
        time.sleep(ACK_RESEND_SLEEP_S)

def try_write(inst, cmd):
    try:
        inst.write(cmd)
        return True
    except Exception:
        return False

def excel_operate(iter_excel):
    """
    返回：(cmd_bytes, info_str)
    """
    try:
        row = next(iter_excel)
        readPhase = int(row[4])
        readwaveA = int(row[5])
        readwaveB = int(row[6])

        bWrite = [0] * 8
        bWrite[0] = 0xFF
        bWrite[1] = 0xFF
        bWrite[2] = (readPhase >> 8) & 0xFF
        bWrite[3] = (readPhase >> 0) & 0xFF
        bWrite[4] = (readwaveA >> 8) & 0xFF
        bWrite[5] = (readwaveA >> 0) & 0xFF
        bWrite[6] = (readwaveB >> 8) & 0xFF
        bWrite[7] = (readwaveB >> 0) & 0xFF

        cmd = bytes(bWrite)
        info = f"phase={readPhase}, waveA={readwaveA}, waveB={readwaveB}"
        return cmd, info
    except Exception:
        return None, None

def parse_arr(reply: str):
    if reply is None:
        return []
    s = str(reply).strip().replace("\r", "").replace("\n", "")
    if not s:
        return []
    out = []
    for p in s.split(","):
        p = p.strip()
        if not p:
            continue
        try:
            out.append(float(p))
        except Exception:
            pass
    return out

def trunc6(x):
    """mW 建议保留 6 位小数更有意义"""
    if x is None or (isinstance(x, float) and math.isnan(x)):
        return "nan"
    try:
        return f"{float(x):.6f}"
    except Exception:
        return "nan"

def trunc3(x):
    if x is None or (isinstance(x, float) and math.isnan(x)):
        return "nan"
    try:
        return f"{math.trunc(float(x) * 1000) / 1000.0:.3f}"
    except Exception:
        return "nan"

def dbm_to_mw(dbm: float) -> float:
    # mW = 10^(dBm/10)
    return 10 ** (dbm / 10.0)

def get_top2_peaks_from_arrays(wav_list, pow_list_dbm):
    n = min(len(wav_list), len(pow_list_dbm))
    wav_list = wav_list[:n]
    pow_list_dbm = pow_list_dbm[:n]

    def is_small_int(x):
        return isinstance(x, (int, float)) and abs(x - round(x)) < 1e-9 and 0 <= x <= 200

    # 去掉头部计数 N
    if n >= 2 and is_small_int(wav_list[0]) and is_small_int(pow_list_dbm[0]):
        w1 = wav_list[1]
        if (500 <= w1 <= 20000) or (1e-9 <= w1 <= 1e-3):
            wav_list = wav_list[1:]
            pow_list_dbm = pow_list_dbm[1:]
            n -= 1

    # 单位判断：m->nm / nm直接用
    wav_nm = []
    for w in wav_list:
        w = float(w)
        if 1e-9 <= w <= 1e-3:
            wav_nm.append(w * 1e9)
        else:
            wav_nm.append(w)

    valid = [i for i in range(n) if 500 <= wav_nm[i] <= 20000]
    if not valid:
        return float("nan"), float("nan"), float("nan"), float("nan")

    valid.sort(key=lambda i: float(pow_list_dbm[i]), reverse=True)

    i1 = valid[0]
    wav1_nm = wav_nm[i1]
    pow1_mw = dbm_to_mw(float(pow_list_dbm[i1]))

    wav2_nm = pow2_mw = float("nan")
    if len(valid) >= 2:
        i2 = valid[1]
        wav2_nm = wav_nm[i2]
        pow2_mw = dbm_to_mw(float(pow_list_dbm[i2]))

    return wav1_nm, pow1_mw, wav2_nm, pow2_mw

def read_two_peaks_stable(inst):
    inst.write(":INIT")
    inst.write("*TRG")
    wav_reply = inst.query(":FETC:ARR:POW:WAV?")
    pow_reply = inst.query(":FETC:ARR:POW?")  # dBm
    wav_list = parse_arr(wav_reply)
    pow_list_dbm = parse_arr(pow_reply)
    return get_top2_peaks_from_arrays(wav_list, pow_list_dbm)

def main():
    excel_path = select_excel_file()
    if not excel_path:
        print("未选择文件，程序退出")
        return

    print(f"已选择文件: {excel_path}")

    wb_in = load_workbook(excel_path, read_only=True, data_only=True)
    ws_in = wb_in.active
    iter_excel = ws_in.iter_rows(min_row=2, values_only=True)

    rm = pyvisa.ResourceManager()
    inst = rm.open_resource(ADDR)
    inst.timeout = VISA_TIMEOUT_MS
    inst.read_termination = "\n"
    inst.write_termination = "\n"

    try:
        try:
            print("Connected:", inst.query("*IDN?").strip())
        except Exception:
            print("提示：*IDN? 不响应也没关系。")

        # 设置单位仍然用 dBm（我们在软件里转 mW）
        try_write(inst, ":INIT:CONT OFF")
        try_write(inst, ":TRIG:SOUR BUS")
        try_write(inst, ":FORM:NDAT 0NM")
        try_write(inst, ":UNIT:POW DBM")

        desktop = get_desktop_path()
        xlsx_path = os.path.join(desktop, XLSX_BASENAME)
        # file_exists = os.path.isfile(xlsx_path)

        # with open(xlsx_path, "a", newline="", encoding="utf-8") as f:
        #     writer = csv.writer(f)
        wb_out = openpyxl.Workbook()
        ws_out = wb_out.active

        head_data = ["timestamp_iso",
                "peak1_wavelength_nm", "peak1_power_mW",
                "peak2_wavelength_nm", "peak2_power_mW"]
        gap_len = len(head_data)+1

        print("开始记录（ACK 不成功则一直重发同一行，不推进 Excel）。Ctrl+C 结束。")

        count = 0
        # since_flush = 0
        # since_row = 2
        loop_time = -1

        # 缓存当前行（ACK 成功后才读取下一行）
        current_cmd = None
        current_info = None
        last_current = True

        while True:
            # 循环写入表头
            if last_current:
                count=0
                loop_time += 1
                iter_excel = ws_in.iter_rows(min_row=2, values_only=True)
                for i, value in enumerate(head_data):
                    ws_out.cell(row=1, column=i+1+loop_time*gap_len, value=value)
                last_current = False
            
            # 只有当当前没有待发送行时，才取 Excel 下一行
            if current_cmd is None:
                cmd, info = excel_operate(iter_excel)
                if cmd is None:
                    last_current = True
                    continue
                current_cmd, current_info = cmd, info

            # 打印 Excel 读取内容（可控频率）
            if PRINT_EXCEL_EVERY > 0 and ((count + 1) % PRINT_EXCEL_EVERY == 0):
                print("----excel----")
                print(current_info)
                print("----excel----")

            # 串口发送
            serial_write(current_cmd)

            # 等 ACK（不成功就一直重发，不推进）
            wait_ack_0x21_forever(current_cmd)

            # ACK 成功：计数并“推进到下一行”
            count += 1
            print("计数:" + str(count))
            current_cmd = None
            current_info = None

            # 读 AQ（失败短重试）
            ts = datetime.now().isoformat(timespec="milliseconds")
            wav1_nm = pow1_mw = wav2_nm = pow2_mw = float("nan")
            for _ in range(VISA_RETRY + 1):
                try:
                    wav1_nm, pow1_mw, wav2_nm, pow2_mw = read_two_peaks_stable(inst)
                    break
                except Exception:
                    time.sleep(0.01)

            # 写 CSV（波长保留3位，功率mW保留6位）
            cur_row = [ts, trunc3(wav1_nm), trunc6(pow1_mw), trunc3(wav2_nm), trunc6(pow2_mw)]
            for i, value in enumerate(cur_row):
                ws_out.cell(row=count+1, column=i+1+loop_time*gap_len, value=value)

    except KeyboardInterrupt:
        print("\n用户中断，已停止。")
        wb_out.save(xlsx_path)

    finally:
        try:
            inst.close()
        finally:
            rm.close()
        wb_out.save(xlsx_path)
        print("结束。")

if __name__ == "__main__":
    main()
