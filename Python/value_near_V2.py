# -*- coding: utf-8 -*-
import os
import sys
import time
import math
from decimal import Decimal, ROUND_FLOOR
from typing import Optional, Tuple, Dict, Any

import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog

from openpyxl import load_workbook, Workbook
from openpyxl.utils import column_index_from_string


def log(msg: str):
    print(msg)
    sys.stdout.flush()


def get_desktop_path():
    home = os.path.expanduser("~")
    for name in ("Desktop", "桌面"):
        p = os.path.join(home, name)
        if os.path.isdir(p):
            return p
    return home


def to_decimal(v) -> Optional[Decimal]:
    """尽量把单元格值转成 Decimal；遇到 NaN/Infinity 返回 None。"""
    if v is None:
        return None

    if isinstance(v, Decimal):
        try:
            if (not v.is_finite()) or v.is_nan():
                return None
        except Exception:
            pass
        return v

    if isinstance(v, (int, float)):
        if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
            return None
        try:
            d = Decimal(str(v))
        except Exception:
            return None
        try:
            if (not d.is_finite()) or d.is_nan():
                return None
        except Exception:
            pass
        return d

    if isinstance(v, str):
        s = v.strip()
        if not s:
            return None
        sl = s.lower()
        if sl in ("nan", "+nan", "-nan", "inf", "+inf", "-inf", "infinity", "+infinity", "-infinity"):
            return None

        # 兼容带逗号/单位：'1,525.00' 或 '1525.00nm'
        s2 = s.replace(",", "")
        for suf in ("nm", "NM"):
            if s2.endswith(suf):
                s2 = s2[:-len(suf)].strip()

        try:
            d = Decimal(s2)
        except Exception:
            return None
        try:
            if (not d.is_finite()) or d.is_nan():
                return None
        except Exception:
            pass
        return d

    return None


def is_literal_nan(v) -> bool:
    """只判断单元格里是否是字母 'nan'（忽略大小写和首尾空格）。"""
    return isinstance(v, str) and v.strip().lower() == "nan"


def mw_to_dbm(mw: Decimal) -> Optional[Decimal]:
    if mw is None or mw <= 0:
        return None
    return Decimal(str(10.0 * math.log10(float(mw))))


def should_skip_by_abs_dbm_delta_lt_threshold(mw1, mw3, threshold_db: Decimal) -> bool:
    """True => 跳过：abs(dBm(mw1) - dBm(mw3)) < threshold_db"""
    m1 = to_decimal(mw1)
    m3 = to_decimal(mw3)
    if m1 is None or m3 is None:
        return False
    d1 = mw_to_dbm(m1)
    d3 = mw_to_dbm(m3)
    if d1 is None or d3 is None:
        return False
    return abs(d1 - d3) < threshold_db


def nearest_even_cent_key_and_dist(x: Decimal) -> Tuple[int, Decimal]:
    """
    目标点key单位：0.01nm 的“分”，且强制偶数 => 0.02nm 网格
    dist 返回单位：nm
    """
    m = x * Decimal("100")
    floor_int = int(m.to_integral_value(rounding=ROUND_FLOOR))
    n_down = floor_int if (floor_int % 2 == 0) else (floor_int - 1)
    n_up = n_down + 2

    dist_down = abs(m - Decimal(n_down))
    dist_up = abs(m - Decimal(n_up))
    n_best = n_down if dist_down <= dist_up else n_up

    dist_cent = abs(m - Decimal(n_best))   # “分”
    dist_nm = dist_cent / Decimal("100")   # 转成 nm
    return n_best, dist_nm


def select_excel_file(root) -> str:
    return filedialog.askopenfilename(
        title="请选择要筛选的Excel文件",
        filetypes=[("Excel文件", "*.xlsx *.xlsm *.xltx *.xltm"), ("所有文件", "*.*")],
        initialdir=get_desktop_path()
    )


def select_save_path(root, default_name: str) -> str:
    return filedialog.asksaveasfilename(
        title="请选择筛选后Excel的保存位置",
        defaultextension=".xlsx",
        initialdir=get_desktop_path(),
        initialfile=default_name,
        filetypes=[("Excel文件", "*.xlsx")]
    )


def _fmt_eta(seconds: float) -> str:
    if seconds < 0 or seconds == float("inf"):
        return "?"
    m, s = divmod(int(seconds), 60)
    h, m = divmod(m, 60)
    if h > 0:
        return f"{h}h{m:02d}m"
    if m > 0:
        return f"{m}m{s:02d}s"
    return f"{s}s"


def pick_best_rows_by_target(
    in_path: str,
    sheet: Optional[str],
    header: bool,
    col_idx: int,
    progress_every_rows: int = 2000,
    diff_threshold_db: Decimal = Decimal("15"),
    diag_range_nm: Tuple[Decimal, Decimal] = (Decimal("1525"), Decimal("1565")),
    compare_col_idx: Optional[int] = None,
    enable_c_filter: bool = True,
    wl_diff_threshold_nm: Decimal = Decimal("0.01"),
    target_half_window_nm: Decimal = Decimal("0.01"),
    b_print_limit: int = 120,
    a_dist_threshold_nm: Decimal = Decimal("0.005"),
) -> Tuple[Dict[int, int], Dict[str, Any]]:
    """
    处理范围：1525-1565（0.02nm网格）
    过滤与筛选：
      2.1 C类过滤（可开关）：abs(wl-other) > wl_diff_threshold_nm => 去除（只计数）
      2.2 B类过滤：abs(dBm(mW1)-dBm(mW3)) < diff_threshold_db => 去除（打印明细）
      2.3 best选择：先要求 dist<=target_half_window_nm
          - 若 enable_c_filter=True：关闭“对比差值最小优先”，按 dist 更小、再光强更大
          - 若 enable_c_filter=False：启用“对比差值最小优先”，
            并增加优先规则：
            先判断 波长列+2 和 波长列+8 是否都为字母 nan；
            若是，则优先；若有多行同样满足，再按：
              1) cmp_diff 更小
              2) intensity 更大
              3) dist 更小
    A类判定：扫描结束后，对每个目标点：best_dist <= a_dist_threshold_nm 才算取到
    """
    wb_val = load_workbook(in_path, read_only=True, data_only=True)
    ws_val = wb_val[sheet] if sheet else wb_val.active

    start_row = 2 if header else 1
    max_row = ws_val.max_row

    range_start_nm, range_end_nm = diag_range_nm
    start_key = int((range_start_nm * Decimal("100")).to_integral_value(rounding=ROUND_FLOOR))
    end_key = int((range_end_nm * Decimal("100")).to_integral_value(rounding=ROUND_FLOOR))
    if start_key % 2 != 0:
        start_key += 1
    if end_key % 2 != 0:
        end_key -= 1

    intensity_col_idx = col_idx + 1
    diff_col1_idx = col_idx + 1
    diff_col3_idx = col_idx + 3

    use_cmp_min_rule = (not enable_c_filter)

    # best:
    #  - use_cmp_min_rule=True  => key -> (nan_both, cmp_diff, intensity, dist_nm, row_idx)
    #  - use_cmp_min_rule=False => key -> (dist_nm, intensity, row_idx)
    best: Dict[int, Any] = {}

    scanned = 0
    invalid_wavelength_rows = 0
    skipped_c_count = 0
    skipped_b_count = 0

    b_removed_records = []  # (row_idx, wl, target_wl, dbm_delta)

    t0 = time.time()
    for row_idx, row_vals in enumerate(ws_val.iter_rows(min_row=start_row, values_only=True), start=start_row):
        scanned += 1

        wl = to_decimal(row_vals[col_idx - 1] if (col_idx - 1) < len(row_vals) else None)
        if wl is None:
            invalid_wavelength_rows += 1
            continue

        key, dist_nm = nearest_even_cent_key_and_dist(wl)
        if not (start_key <= key <= end_key):
            continue

        # 2.1 C类过滤
        if enable_c_filter and (compare_col_idx is not None):
            other_raw = row_vals[compare_col_idx - 1] if (compare_col_idx - 1) < len(row_vals) else None
            other = to_decimal(other_raw)
            if other is not None and abs(wl - other) > wl_diff_threshold_nm:
                skipped_c_count += 1
                continue

        # 2.2 B类过滤
        mw1 = row_vals[diff_col1_idx - 1] if (diff_col1_idx - 1) < len(row_vals) else None
        mw3 = row_vals[diff_col3_idx - 1] if (diff_col3_idx - 1) < len(row_vals) else None

        m1 = to_decimal(mw1)
        m3 = to_decimal(mw3)
        dbm_delta = None
        if m1 is not None and m3 is not None:
            d1 = mw_to_dbm(m1)
            d3 = mw_to_dbm(m3)
            if d1 is not None and d3 is not None:
                dbm_delta = abs(d1 - d3)

        if should_skip_by_abs_dbm_delta_lt_threshold(mw1, mw3, diff_threshold_db):
            skipped_b_count += 1
            if len(b_removed_records) < b_print_limit:
                target_wl = Decimal(key) / Decimal("100")
                b_removed_records.append((row_idx, wl, target_wl, dbm_delta))
            continue

        # 2.3 best选择：先卡窗口
        if dist_nm > target_half_window_nm:
            continue

        intensity = to_decimal(row_vals[intensity_col_idx - 1] if (intensity_col_idx - 1) < len(row_vals) else None)
        intensity = intensity if intensity is not None else Decimal("-Infinity")

        if use_cmp_min_rule:
            # 波长列+2 和 波长列+8 都是字母 nan，则优先
            v_plus2 = row_vals[(col_idx + 2) - 1] if (col_idx + 2 - 1) < len(row_vals) else None
            v_plus8 = row_vals[(col_idx + 8) - 1] if (col_idx + 8 - 1) < len(row_vals) else None
            nan_both = is_literal_nan(v_plus2) and is_literal_nan(v_plus8)

            if compare_col_idx is not None:
                other_raw = row_vals[compare_col_idx - 1] if (compare_col_idx - 1) < len(row_vals) else None
                other = to_decimal(other_raw)
                cmp_diff = abs(wl - other) if other is not None else Decimal("Infinity")
            else:
                cmp_diff = Decimal("Infinity")

            # 比较顺序：
            # 0) nan_both=True 优先
            # 1) cmp_diff 越小越好
            # 2) intensity 越大越好
            # 3) dist_nm 越小越好
            cand = (nan_both, cmp_diff, intensity, dist_nm, row_idx)

            if key not in best:
                best[key] = cand
            else:
                best_nan, best_cmp, best_inten, best_dist, _ = best[key]
                better = False

                if nan_both and (not best_nan):
                    better = True
                elif nan_both == best_nan:
                    if cmp_diff < best_cmp:
                        better = True
                    elif cmp_diff == best_cmp:
                        if intensity > best_inten:
                            better = True
                        elif intensity == best_inten:
                            if dist_nm < best_dist:
                                better = True

                if better:
                    best[key] = cand

        else:
            # C类过滤开启时：只比 dist，再比 intensity
            cand = (dist_nm, intensity, row_idx)
            if key not in best:
                best[key] = cand
            else:
                best_dist, best_inten, _ = best[key]
                if dist_nm < best_dist or (dist_nm == best_dist and intensity > best_inten):
                    best[key] = cand

        if scanned % progress_every_rows == 0 or row_idx == max_row:
            now = time.time()
            elapsed = now - t0
            speed = scanned / elapsed if elapsed > 0 else 0.0
            remaining_rows = (max_row - row_idx)
            eta = remaining_rows / speed if speed > 0 else float("inf")
            log(
                f"[过滤+筛选] 行 {row_idx}/{max_row} ({row_idx/max_row*100:.1f}%) | "
                f"目标点数={len(best)} | C去除={skipped_c_count} | B去除={skipped_b_count} | "
                f"波长无效行={invalid_wavelength_rows} | {speed:.0f} 行/s | ETA {_fmt_eta(eta)}"
            )

    if not best:
        raise RuntimeError("过滤后未找到可用数据（可能阈值过严、窗口过小或列选择不对）。")

    if use_cmp_min_rule:
        best_map = {k: r for k, (_, _, _, _, r) in best.items()}
        best_dist_by_key = {k: d for k, (_, _, _, d, _) in best.items()}
    else:
        best_map = {k: r for k, (_, _, r) in best.items()}
        best_dist_by_key = {k: d for k, (d, _, _) in best.items()}

    diag = {
        "range_nm": (range_start_nm, range_end_nm),
        "range_key": (start_key, end_key),

        "scanned_rows_total": scanned,
        "invalid_wavelength_rows": invalid_wavelength_rows,

        "compare_col_idx": compare_col_idx,
        "enable_c_filter": enable_c_filter,
        "wl_diff_threshold_nm": wl_diff_threshold_nm,
        "target_half_window_nm": target_half_window_nm,

        "diff_threshold_db": diff_threshold_db,
        "skipped_c_count": skipped_c_count,
        "skipped_b_count": skipped_b_count,

        "b_removed_records": b_removed_records,
        "b_print_limit": b_print_limit,

        "best_dist_by_key": best_dist_by_key,
        "a_dist_threshold_nm": a_dist_threshold_nm,
    }
    return best_map, diag


def copy_kept_rows_valueonly_fast(
    in_path: str,
    out_path: str,
    sheet: Optional[str],
    kept_rows_sorted: list,
    header: bool,
    progress_every_rows: int = 200
):
    """
    只保留值和列位置，不复制样式。
    使用流式读取 + 流式写出，加快速度。
    """
    wb_in = load_workbook(in_path, data_only=True, read_only=True)
    ws_in = wb_in[sheet] if sheet else wb_in.active

    wb_out = Workbook(write_only=True)
    ws_out = wb_out.create_sheet(title=(ws_in.title[:31] if ws_in.title else "Sheet1"))

    rows_to_copy = []
    if header:
        rows_to_copy.append(1)
    rows_to_copy.extend(kept_rows_sorted)

    if not rows_to_copy:
        if os.path.exists(out_path):
            try:
                os.remove(out_path)
            except PermissionError:
                pass
        wb_out.save(out_path)
        return

    total = len(rows_to_copy)
    copied = 0
    p = 0
    max_needed_row = rows_to_copy[-1]

    t0 = time.time()

    for row_idx, row_vals in enumerate(
        ws_in.iter_rows(min_row=1, max_row=max_needed_row, values_only=True),
        start=1
    ):
        if p >= total:
            break

        target_row = rows_to_copy[p]

        if row_idx < target_row:
            continue
        elif row_idx == target_row:
            ws_out.append(list(row_vals))
            copied += 1
            p += 1

            if copied % progress_every_rows == 0 or copied == total:
                elapsed = time.time() - t0
                speed = copied / elapsed if elapsed > 0 else 0.0
                eta = (total - copied) / speed if speed > 0 else float("inf")
                log(f"[复制] {copied}/{total} ({copied/total*100:.1f}%) | {speed:.0f} 行/s | ETA {_fmt_eta(eta)}")

    if os.path.exists(out_path):
        try:
            os.remove(out_path)
        except PermissionError:
            pass

    wb_out.save(out_path)


def _parse_decimal_or_default(s: str, default: Decimal) -> Decimal:
    if s is None:
        return default
    ss = str(s).strip()
    if ss == "":
        return default
    try:
        return Decimal(ss)
    except Exception:
        return default


def _print_wl_list(title: str, wls: list, per_line: int = 20):
    log(f"{title} 数量：{len(wls)}")
    if not wls:
        return
    line = []
    for i, wl in enumerate(wls, 1):
        line.append(f"{wl:.2f}")
        if i % per_line == 0:
            log("  " + ", ".join(line))
            line = []
    if line:
        log("  " + ", ".join(line))


def _print_missing_grouped(best_map: Dict[int, int], diag: Dict[str, Any]):
    (start_nm, end_nm) = diag["range_nm"]
    (start_key, end_key) = diag["range_key"]

    log(f"==== 输出（{start_nm}-{end_nm}，步长0.02nm）====")
    log(f"总扫描行：{diag.get('scanned_rows_total', 0)} | 波长无效行：{diag.get('invalid_wavelength_rows', 0)}")
    log(f"目标点窗口：±{diag.get('target_half_window_nm', Decimal('0.01'))} nm")
    log("--------------------------------------------------")

    best_dist_by_key: Dict[int, Decimal] = diag.get("best_dist_by_key", {})
    a_thr: Decimal = diag.get("a_dist_threshold_nm", Decimal("0.005"))

    missing_a = []
    for key in range(start_key, end_key + 1, 2):
        d = best_dist_by_key.get(key)
        if d is None or d > a_thr:
            missing_a.append(Decimal(key) / Decimal("100"))

    _print_wl_list(f"A类未取到目标点（2.3后判定：best_dist <= {a_thr}nm 才算取到）", missing_a, per_line=20)

    log("--------------------------------------------------")
    if not diag.get("enable_c_filter", True):
        log("C类过滤：关闭")
    else:
        thr = diag.get("wl_diff_threshold_nm", Decimal("0.01"))
        log(f"C类过滤：开启 | 条件 abs(波长-对比列) > {thr} nm")
        log(f"C类去除行数：{diag.get('skipped_c_count', 0)}")

    log("--------------------------------------------------")
    b_cnt = diag.get("skipped_b_count", 0)
    thr_db = diag.get("diff_threshold_db", Decimal("15"))
    log(f"B类去除行数（abs(dBm(mW1)-dBm(mW3)) < {thr_db} dB）：{b_cnt}")

    recs = diag.get("b_removed_records", [])
    limit = diag.get("b_print_limit", 120)
    if b_cnt > 0 and limit != 0:
        if b_cnt > len(recs):
            log(f"仅打印前 {len(recs)} 条B类去除明细（上限 {limit} 条），其余省略。")
        if recs:
            log("B类去除明细：row | 波长 | 目标点 | dBm差值(abs)")
            for (row_idx, wl, target_wl, dbm_delta) in recs:
                d_str = "None" if dbm_delta is None else str(dbm_delta)
                log(f"  {row_idx} | {wl} | {target_wl} | {d_str}")

    log("--------------------------------------------------")
    log(f"最终保留目标点数量：{len(best_map)}")
    log("==== 输出结束 ====")


def main():
    root = tk.Tk()
    root.withdraw()
    try:
        root.attributes("-topmost", True)
    except Exception:
        pass

    in_path = select_excel_file(root)
    if not in_path:
        root.destroy()
        return

    ext = os.path.splitext(in_path)[1].lower()
    if ext == ".xls":
        messagebox.showerror("格式不支持", "openpyxl 不支持 .xls，请先另存为 .xlsx 或 .xlsm。")
        root.destroy()
        return

    base = os.path.splitext(os.path.basename(in_path))[0]
    out_path = select_save_path(root, f"{base}_filtered.xlsx")
    if not out_path:
        root.destroy()
        return

    header = messagebox.askyesno("表头设置", "第1行是否是表头？\n（是：保留第1行）")

    sheet = simpledialog.askstring("工作表名称", "请输入工作表名称（留空=使用当前激活表）：")
    sheet = (sheet.strip() if sheet else "")
    sheet = sheet if sheet else None

    col_letter = simpledialog.askstring("列选择", "请输入要处理的列字母（默认 I；例如 L）：")
    col_letter = (col_letter.strip().upper() if col_letter else "I")
    try:
        col_idx = column_index_from_string(col_letter)
    except Exception:
        messagebox.showerror("列输入错误", f"列字母无效：{col_letter}")
        root.destroy()
        return

    compare_col_letter = simpledialog.askstring(
        "对比列选择",
        "请输入用于对比的列字母（留空=不使用对比列）：\n例如 J / K",
    )
    compare_col_letter = (compare_col_letter.strip().upper() if compare_col_letter else "")
    if compare_col_letter:
        try:
            compare_col_idx = column_index_from_string(compare_col_letter)
        except Exception:
            messagebox.showerror("列输入错误", f"对比列字母无效：{compare_col_letter}")
            root.destroy()
            return
    else:
        compare_col_idx = None

    enable_c_filter = messagebox.askyesno(
        "C类过滤",
        "是否启用C类过滤？\n\n"
        "【是】= 开启C类过滤（abs(波长-对比列) > 阈值则剔除），并关闭“差值最小优先”规则\n"
        "【否】= 关闭C类过滤，启用“差值最小优先”规则（需要对比列）"
    )

    if enable_c_filter:
        wl_thr_str = simpledialog.askstring(
            "C类阈值(nm)",
            "请输入C类阈值(nm)：\n当 abs(波长列-对比列) > 阈值 时，该行跳过。\n\n留空默认 0.01",
        )
        wl_diff_threshold_nm = _parse_decimal_or_default(wl_thr_str, Decimal("0.01"))
    else:
        wl_diff_threshold_nm = Decimal("0.01")

    win_str = simpledialog.askstring(
        "目标点窗口(±nm)",
        "请输入目标点窗口大小（±nm）：\n"
        "只有落在【目标点 ± 窗口】内的行才参与该目标点best竞争。\n\n"
        "留空默认 0.01",
    )
    target_half_window_nm = _parse_decimal_or_default(win_str, Decimal("0.01"))

    thr_str = simpledialog.askstring(
        "B类阈值(dB)",
        "请输入B类阈值(dB)：\n当 abs(dBm(mW1) - dBm(mW3)) < 阈值 时，该行跳过不参与筛选。\n\n留空默认 15",
    )
    diff_threshold_db = _parse_decimal_or_default(thr_str, Decimal("15"))

    b_limit_str = simpledialog.askstring(
        "B类打印条数上限",
        "请输入B类去除明细打印条数上限（留空默认120；0=不打印明细只打印数量）："
    )
    try:
        b_print_limit = int(str(b_limit_str).strip()) if (b_limit_str is not None and str(b_limit_str).strip() != "") else 120
        if b_print_limit < 0:
            b_print_limit = 120
    except Exception:
        b_print_limit = 120

    a_thr_str = simpledialog.askstring(
        "A类接近阈值(nm)",
        "请输入A类“取到目标点”的接近阈值(nm)：\n"
        "当最终 best_dist <= 该阈值 时，认为该目标点已取到，不输出。\n"
        "留空默认 0.005（较严格）；\n"
        "如果想“只要有best就算取到”，可填 0.01；想严格一致填 0。",
    )
    a_dist_threshold_nm = _parse_decimal_or_default(a_thr_str, Decimal("0.005"))

    if (not enable_c_filter) and (compare_col_idx is None):
        messagebox.showwarning(
            "提示",
            "你选择了关闭C类过滤（将启用“差值最小优先”规则），但未提供对比列。\n"
            "将自动退化为：先看 nan 优先规则，再按光强和距离比较。"
        )

    scan_every = 2000
    copy_every = 200

    try:
        log("==== 开始扫描 ====")
        log(f"范围：1525-1565 | 目标点窗口：±{target_half_window_nm} nm")
        log(f"波长列：{col_letter}")
        log(f"对比列：{compare_col_letter if compare_col_idx is not None else '未使用'}")
        log(f"C类过滤：{'开启' if enable_c_filter else '关闭'}")
        if enable_c_filter:
            log(f"  C类阈值：abs({col_letter}-{compare_col_letter}) > {wl_diff_threshold_nm} nm（只计数）")
        else:
            log("  C类过滤关闭：启用 nan 优先规则 + 差值最小优先规则")
            log(f"  nan 优先规则：{col_letter}+2 和 {col_letter}+8 都为字母 nan 时优先")
            log("  后续比较顺序：1) compare差值更小 2) 强度更大 3) 离目标点更近")
        log(f"B类阈值：abs(dBm(mW1)-dBm(mW3)) < {diff_threshold_db} dB（打印明细）")
        log(f"A类判定阈值：best_dist <= {a_dist_threshold_nm} nm 才算取到目标点")
        log(f"B类明细打印上限：{b_print_limit}")

        best_map, diag = pick_best_rows_by_target(
            in_path, sheet, header, col_idx,
            progress_every_rows=scan_every,
            diff_threshold_db=diff_threshold_db,
            diag_range_nm=(Decimal("1525"), Decimal("1565")),
            compare_col_idx=compare_col_idx,
            enable_c_filter=enable_c_filter,
            wl_diff_threshold_nm=wl_diff_threshold_nm,
            target_half_window_nm=target_half_window_nm,
            b_print_limit=b_print_limit,
            a_dist_threshold_nm=a_dist_threshold_nm,
        )

        _print_missing_grouped(best_map, diag)

        kept_rows = sorted(best_map.values())
        log(f"==== 扫描完成：目标点数量={len(best_map)}，保留行数={len(kept_rows)}（不含表头） ====")

        log("==== 开始复制：只保留值和列位置，不复制样式 ====")
        copy_kept_rows_valueonly_fast(in_path, out_path, sheet, kept_rows, header, progress_every_rows=copy_every)

        messagebox.showinfo(
            "完成",
            f"已输出：\n{out_path}\n\n"
            f"目标点数量：{len(best_map)}\n"
            f"保留数据行数：{len(kept_rows)}（不含表头）\n"
            f"波长列：{col_letter}\n"
            f"对比列：{compare_col_letter if compare_col_idx is not None else '未使用'}\n"
            f"C类过滤：{'开启' if enable_c_filter else '关闭'}\n"
            f"C类去除行数：{diag.get('skipped_c_count', 0)}\n"
            f"B类去除行数：{diag.get('skipped_b_count', 0)}\n"
            f"目标点窗口：±{target_half_window_nm} nm\n"
            f"A类判定阈值：best_dist <= {a_dist_threshold_nm} nm"
        )
        log("完成，输出：" + out_path)

    except PermissionError as e:
        messagebox.showerror(
            "失败：无写入权限",
            f"{e}\n\n常见原因：输出文件正在被Excel/WPS打开，或无写权限。\n"
            f"请关闭输出文件/结束EXCEL进程，或更换输出目录（如 C:\\Temp）后重试。"
        )
        raise
    except Exception as e:
        messagebox.showerror("失败", str(e))
        raise
    finally:
        root.destroy()


if __name__ == "__main__":
    main()